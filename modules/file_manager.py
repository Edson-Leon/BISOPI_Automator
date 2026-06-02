# ═══════════════════════════════════════════════════════════
# BISOPI Automator v1.0 — Mayo 2026
# Autor: Edson Leon
# Cargo: Data Insights Lead — Operaciones
# Empresa: Bision Consulting
# Contacto: edson.leon@bisionconsulting.com
# ═══════════════════════════════════════════════════════════

"""
Abstracción para apertura y guardado del archivo de plantilla.

Soporta dos modos:
  - Ruta local (PLANTILLA_PATH configurado) : lee/escribe directamente en disco.
  - Modo descarga                           : serializa a bytes para st.download_button.
"""
from __future__ import annotations

import io

import streamlit as st
from openpyxl import load_workbook, Workbook


def check_writable(filepath: str) -> bool:
    """
    Devuelve True si el archivo puede abrirse para escritura (no está bloqueado).
    Devuelve False si está abierto en Excel u otro proceso.

    No muestra ningún mensaje — el caller decide qué mostrar.
    """
    if not filepath:
        return True
    try:
        with open(filepath, "r+b"):
            pass
        return True
    except PermissionError:
        return False
    except OSError:
        return False


def get_plantilla(source=None) -> tuple[Workbook | None, str | None]:
    """
    Abre el workbook de la plantilla.

    Parámetros
    ----------
    source : str | UploadedFile | None
        str          → ruta local en disco.
        UploadedFile → objeto de st.file_uploader.
        None         → devuelve (None, None).

    Retorna
    -------
    (wb, effective_path)
        wb             : Workbook abierto, o None si source es None.
        effective_path : ruta en disco si viene de archivo local;
                         None si viene de upload.
    """
    if source is None:
        return None, None

    if isinstance(source, str):
        wb = load_workbook(source, keep_vba=False, data_only=False)
        return wb, source

    # Objeto UploadedFile de Streamlit (o cualquier objeto con .getvalue())
    raw = source.getvalue()
    wb  = load_workbook(io.BytesIO(raw), keep_vba=False, data_only=False)
    return wb, None


def save_plantilla(
    wb: Workbook,
    filepath: str | None,
) -> tuple[bool, bytes | None]:
    """
    Guarda el workbook.

    Comportamiento según `filepath`:
      - No None → guarda en disco y retorna (True, None).
      - None    → serializa a bytes y retorna (False, bytes).

    En ambos casos: si ocurre PermissionError u otro error, muestra
    st.warning y retorna (False, None).

    Parámetros
    ----------
    wb       : Workbook ya modificado.
    filepath : Ruta destino en disco, o None para modo descarga.

    Retorna
    -------
    (saved_to_disk, bytes_or_none)
        saved_to_disk : True solo cuando se guardó en disco correctamente.
        bytes_or_none : bytes del xlsx cuando filepath es None y no hubo error;
                        None en todos los demás casos.
    """
    # ── Metadatos de autoría ──────────────────────────────────────────────────
    from datetime import datetime as _dt
    wb.properties.creator        = "Edson Leon"
    wb.properties.lastModifiedBy = "Edson Leon"
    wb.properties.company        = "Bision Consulting"
    wb.properties.description    = (
        "BISOPI Automator v1.0 — Mayo 2026. "
        "Herramienta de automatización para el registro de horas en BISOPI. "
        "Arquitectura escalable hacia integración con Azure DevOps y calendario de Outlook."
    )
    wb.properties.version        = "1.0"
    # ─────────────────────────────────────────────────────────────────────────

    if filepath:
        try:
            wb.save(filepath)
            return True, None
        except PermissionError:
            st.warning(
                "⚠️ **No se pudo guardar el archivo** — "
                "ciérralo en Excel e intenta de nuevo. "
                "Los cambios no se perdieron; recarga la app para reintentar."
            )
            return False, None
        except Exception as exc:
            st.warning(f"⚠️ Error al guardar el archivo: {exc}")
            return False, None
    else:
        try:
            buf = io.BytesIO()
            wb.save(buf)
            return False, buf.getvalue()
        except Exception as exc:
            st.warning(f"⚠️ Error al serializar el archivo: {exc}")
            return False, None
