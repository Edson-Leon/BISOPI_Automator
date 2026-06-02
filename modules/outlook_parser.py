# ═══════════════════════════════════════════════════════════
# BISOPI Automator v1.0 — Mayo 2026
# Autor: Edson Leon
# Cargo: Data Insights Lead — Operaciones
# Empresa: Bision Consulting
# Contacto: edson.leon@bisionconsulting.com
# ═══════════════════════════════════════════════════════════

"""
Parsea archivos .ics exportados de Outlook y los convierte en
DataFrames compatibles con REGISTRO_COLUMNS.

Todos los eventos del rango de fechas se incluyen en el DataFrame:

  [BISOPI] + descripción completa  → Estado = "Pendiente"
  [BISOPI] + faltan campos clave   → Estado = "⚠ Incompleto"
  Sin [BISOPI]                     → Estado = "⚠ Sin clasificar"
  Sin [BISOPI] + solapado          → Estado = "⚠ Sin clasificar"
                                     RespuestaAPI = "⚠ Solapado con evento BISOPI"

La descripción del evento puede incluir campos clave:valor
para pre-rellenar Proyecto, GrupoTarea, Tarea, TipoHora y Comentario.
"""
from __future__ import annotations

import io
import os
import re
from datetime import date, datetime, timedelta
from typing import IO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill

try:
    from icalendar import Calendar
    _ICAL_AVAILABLE = True
except ImportError:
    _ICAL_AVAILABLE = False

from modules.loader import REGISTRO_COLUMNS


# ── Constantes ────────────────────────────────────────────────────────────────

_BISOPI_TAG = re.compile(r"\[BISOPI\]", re.IGNORECASE)

# Claves reconocidas en la descripción del evento → campo interno del modelo
_KEY_ALIASES: dict[str, str] = {
    "proyecto":    "Proyecto",
    "grupotarea":  "GrupoTarea",
    "grupo tarea": "GrupoTarea",
    "tarea":       "Tarea",
    "tipohora":    "TipoHora",
    "tipo hora":   "TipoHora",
    "comentario":  "Comentario",
}

# Minutos redondeo (hacia abajo al múltiplo de 15)
_ALLDAY_MINUTES = 8 * 60  # eventos de día completo → 8 h

# Relleno amarillo para celdas que el usuario debe completar
_YELLOW_FILL = PatternFill(start_color="FFF9C3", end_color="FFF9C3", fill_type="solid")

# Columnas que se resaltan en amarillo cuando están vacías en el archivo exportado
_HIGHLIGHT_COLS: frozenset = frozenset({"Proyecto", "Grupo Tarea", "Tarea", "Tipo Hora"})

# Mapeo nombre interno del modelo → nombre de columna en la hoja Registro del Excel
_INTERNAL_TO_EXCEL: dict[str, str] = {
    "Proyecto":      "Proyecto",
    "GrupoTarea":    "Grupo Tarea",
    "Tarea":         "Tarea",
    "TipoHora":      "Tipo Hora",
    "FechaRegistro": "Fecha Registro",
    "Horas":         "Horas",
    "Minutos":       "Minutos",
    "HoraInicio":    "Hora Inicio",
    "Comentario":    "Comentario",
    "Estado":        "Estado",
    "RespuestaAPI":  "Respuesta API",
}

# Fila 1 = título, fila 2 = encabezados, fila 3+ = datos (mismo esquema que loader.py)
_REGISTRO_HEADER_ROW = 2
_REGISTRO_DATA_START = 3


# ── Helpers internos ──────────────────────────────────────────────────────────

def _to_naive_local(dt) -> datetime | None:
    """
    Convierte un valor DTSTART/DTEND de icalendar a datetime naive local.
    Maneja datetime con zona horaria, datetime sin zona y date (todo-el-día).
    """
    if dt is None:
        return None
    if isinstance(dt, datetime):
        if dt.tzinfo is not None:
            dt = dt.astimezone().replace(tzinfo=None)
        return dt
    if isinstance(dt, date):
        # Evento de día completo → medianoche del día
        return datetime(dt.year, dt.month, dt.day)
    return None


def _round_down_15(total_minutes: int) -> tuple[int, int]:
    """
    Redondea los minutos totales hacia abajo al múltiplo de 15 más cercano.
    Devuelve (horas, minutos).
    """
    total = max(0, total_minutes)
    mins  = (total % 60) // 15 * 15
    horas = total // 60
    return horas, mins


# ── Función 1: parsear la descripción del evento ──────────────────────────────

def parse_description(description: str) -> dict:
    """
    Extrae campos BISOPI desde la descripción de un evento de Outlook.

    Formato esperado (una clave por línea):
        Proyecto: Nombre del proyecto
        Grupo Tarea: Nombre del grupo
        Tarea: Descripción de la tarea
        Tipo Hora: Laboral | Adicional
        Comentario: Texto libre

    Retorna un dict con las claves internas del modelo
    (Proyecto, GrupoTarea, Tarea, TipoHora, Comentario).
    Las claves no encontradas no aparecen en el dict.
    """
    result: dict = {}
    if not description:
        return result

    for line in description.splitlines():
        if ":" not in line:
            continue
        raw_key, _, raw_val = line.partition(":")
        key_norm = raw_key.strip().lower()
        internal = _KEY_ALIASES.get(key_norm)
        if internal and raw_val.strip():
            result[internal] = raw_val.strip()

    return result


# ── Función interna: pasada 2 — construir DataFrame desde all_evts ───────────

def _process_event_list(all_evts: list[dict]) -> pd.DataFrame:
    """
    Construye el DataFrame BISOPI a partir de una lista normalizada de eventos.

    Recibe la salida de la «pasada 1» (producida por parse_ics o events_to_dataframe)
    y aplica la lógica de clasificación, detección de solapamientos y ordenación.

    Schema esperado de cada dict en all_evts
    ─────────────────────────────────────────
    summary         str       Título del evento
    is_bisopi       bool      True si el título contiene [BISOPI]
    is_allday       bool      True para eventos de día completo
    dtstart         datetime  Inicio naive local
    dtend_eff       datetime  Fin efectivo (dtend real o calculado desde duración)
    event_date      date      Fecha del evento
    horas           int       Horas redondeadas hacia abajo al múltiplo de 15 min
    minutos         int       Minutos residuales (0, 15, 30 o 45)
    description_raw str       Descripción/body en texto plano

    Retorna
    -------
    DataFrame con REGISTRO_COLUMNS, ordenado por fecha y tipo de evento.
    Puede estar vacío si all_evts es vacío o todos los eventos producen filas inválidas.
    """
    if not all_evts:
        return pd.DataFrame(columns=REGISTRO_COLUMNS)

    # Intervalos de eventos BISOPI con hora real (se usan para detección de solapamiento).
    # Se excluyen los eventos de día completo: su intervalo artificial cubriría todo el día.
    bisopi_intervals: list[tuple[datetime, datetime]] = [
        (e["dtstart"], e["dtend_eff"])
        for e in all_evts
        if e["is_bisopi"] and not e["is_allday"]
    ]

    def _overlaps_bisopi(evt: dict) -> bool:
        """True si el intervalo del evento se solapa con algún evento BISOPI timed."""
        if evt["is_allday"]:
            return False
        a_s, a_e = evt["dtstart"], evt["dtend_eff"]
        return any(a_s < b_e and b_s < a_e for b_s, b_e in bisopi_intervals)

    rows: list[dict] = []

    for evt in all_evts:
        fecha_str        = evt["event_date"].strftime("%Y-%m-%d")
        horas            = evt["horas"]
        minutos          = evt["minutos"]
        is_allday        = evt["is_allday"]
        dtstart          = evt["dtstart"]
        hora_inicio_base = "" if is_allday else dtstart.strftime("%H:%M")

        if evt["is_bisopi"]:
            # ── Evento [BISOPI] ───────────────────────────────────────────────
            fields = parse_description(evt["description_raw"])

            tipo = fields.get("TipoHora", "Laboral").strip()
            if tipo not in ("Laboral", "Adicional"):
                tipo = "Laboral"

            # Tarea: descripción → fallback al SUMMARY sin el tag
            tarea_fallback = _BISOPI_TAG.sub("", evt["summary"]).strip(" -–:").strip()
            tarea          = fields.get("Tarea") or tarea_fallback

            # HoraInicio solo para Adicional con hora real
            hora_inicio = hora_inicio_base if tipo == "Adicional" else ""

            # ¿Faltan campos clave?
            missing: list[str] = []
            if not fields.get("Proyecto"):
                missing.append("Proyecto")
            if not fields.get("GrupoTarea"):
                missing.append("Grupo Tarea")

            if missing:
                estado    = "⚠ Incompleto"
                respuesta = "Faltan campos: " + ", ".join(missing)
            else:
                estado    = "Pendiente"
                respuesta = ""

            rows.append({
                "Proyecto":      fields.get("Proyecto",   ""),
                "GrupoTarea":    fields.get("GrupoTarea",  ""),
                "Tarea":         tarea,
                "TipoHora":      tipo,
                "FechaRegistro": fecha_str,
                "Horas":         horas,
                "Minutos":       minutos,
                "HoraInicio":    hora_inicio,
                "Comentario":    fields.get("Comentario", ""),
                "Estado":        estado,
                "RespuestaAPI":  respuesta,
            })

        else:
            # ── Evento sin [BISOPI] ───────────────────────────────────────────
            respuesta = "⚠ Solapado con evento BISOPI" if _overlaps_bisopi(evt) else ""

            rows.append({
                "Proyecto":      "",
                "GrupoTarea":    "",
                "Tarea":         evt["summary"].strip(),
                "TipoHora":      "",
                "FechaRegistro": fecha_str,
                "Horas":         horas,
                "Minutos":       minutos,
                "HoraInicio":    hora_inicio_base,
                "Comentario":    "",
                "Estado":        "⚠ Sin clasificar",
                "RespuestaAPI":  respuesta,
            })

    if not rows:
        return pd.DataFrame(columns=REGISTRO_COLUMNS)

    df = pd.DataFrame(rows, columns=REGISTRO_COLUMNS)

    # Ordenar: fecha → BISOPI primero → Laborales antes que Adicionales
    df["_bisopi"] = (df["Estado"] != "⚠ Sin clasificar").astype(int)
    df["_orden"]  = df["TipoHora"].map({"Laboral": 0, "Adicional": 1}).fillna(2)
    df = (
        df.sort_values(
            ["FechaRegistro", "_bisopi", "_orden"],
            ascending=[True, False, True],
        )
        .drop(columns=["_bisopi", "_orden"])
        .reset_index(drop=True)
    )
    return df


# ── Función 2: parsear archivo ICS ────────────────────────────────────────────

def parse_ics(
    file: IO[bytes],
    week_start: date,
    week_end: date,
) -> pd.DataFrame:
    """
    Lee un archivo .ics y procesa TODOS los eventos del rango
    [week_start, week_end] (inclusive).

    Comportamiento por tipo de evento
    ----------------------------------
    [BISOPI] + descripción completa → Estado = "Pendiente"
    [BISOPI] + faltan Proyecto/GrupoTarea → Estado = "⚠ Incompleto",
        RespuestaAPI lista los campos faltantes.
    Sin [BISOPI] → Estado = "⚠ Sin clasificar",
        solo fecha/hora inicio/duración/título precargados.
    Sin [BISOPI] + solapado con evento BISOPI (mismo intervalo de tiempo) →
        igual al anterior pero RespuestaAPI = "⚠ Solapado con evento BISOPI".

    Parámetros
    ----------
    file        : objeto legible (bytes), p.ej. st.file_uploader o io.BytesIO.
    week_start  : primer día de la semana (inclusive).
    week_end    : último día de la semana (inclusive).

    Retorna
    -------
    DataFrame con REGISTRO_COLUMNS. Puede estar vacío si no hay
    eventos en el rango.

    Lanza
    -----
    ImportError  — si la librería 'icalendar' no está instalada.
    ValueError   — si el archivo no es un ICS válido.
    """
    if not _ICAL_AVAILABLE:
        raise ImportError(
            "La librería 'icalendar' no está instalada. "
            "Ejecuta: pip install icalendar"
        )

    try:
        raw = file.read() if hasattr(file, "read") else file
        cal = Calendar.from_ical(raw)
    except Exception as exc:
        raise ValueError(f"No se pudo parsear el archivo ICS: {exc}") from exc

    # ── Pasada 1: recopilar todos los VEVENT del rango ────────────────────────
    # Se acumulan como dicts con la info necesaria para la pasada 2.

    all_evts: list[dict] = []

    for component in cal.walk():
        if component.name != "VEVENT":
            continue

        summary = str(component.get("SUMMARY", ""))

        _ds = component.get("DTSTART")
        _de = component.get("DTEND")
        _dr = component.get("DURATION")

        dtstart = _to_naive_local(_ds.dt) if _ds is not None else None
        if dtstart is None:
            continue

        if _de is not None:
            dtend = _to_naive_local(_de.dt)
        elif _dr is not None:
            dtend = dtstart + _dr.dt
        else:
            dtend = None

        # Filtrar por semana
        event_date = dtstart.date()
        if not (week_start <= event_date <= week_end):
            continue

        # Detectar evento de día completo (DTSTART es date, no datetime)
        is_allday = not isinstance(_ds.dt, datetime)

        if is_allday:
            total_minutes = _ALLDAY_MINUTES
        elif dtend is not None and dtend > dtstart:
            total_minutes = int((dtend - dtstart).total_seconds() / 60)
        else:
            total_minutes = 60  # fallback: 1 hora

        horas, minutos = _round_down_15(total_minutes)
        if horas == 0 and minutos == 0:
            continue  # duración demasiado corta para registrar

        # dtend efectivo para detección de solapamientos
        dtend_eff = dtend if (dtend is not None and dtend > dtstart) else (
            dtstart + timedelta(hours=horas, minutes=minutos)
        )

        all_evts.append({
            "summary":         summary,
            "is_bisopi":       bool(_BISOPI_TAG.search(summary)),
            "is_allday":       is_allday,
            "dtstart":         dtstart,
            "dtend_eff":       dtend_eff,
            "event_date":      event_date,
            "horas":           horas,
            "minutos":         minutos,
            "description_raw": str(component.get("DESCRIPTION", "")),
        })

    # Delegar en la función compartida (también usada por graph_client.events_to_dataframe)
    return _process_event_list(all_evts)


# ── Función 3: calcular horas laborales faltantes por día ────────────────────

def calculate_gaps(
    df: pd.DataFrame,
    work_hours_per_day: int = 8,
) -> pd.DataFrame:
    """
    Calcula el tiempo laboral no cubierto por día.

    Para cada día con al menos un registro 'Laboral', comprueba si el total
    de minutos laborales es inferior a work_hours_per_day * 60.

    Parámetros
    ----------
    df                  : DataFrame con columnas del modelo (REGISTRO_COLUMNS).
    work_hours_per_day  : horas laborales objetivo por día (default 8).

    Retorna
    -------
    DataFrame con columnas:
        FechaRegistro, DiaSemana, MinutosRegistrados, MinutosFaltantes
    Solo incluye días con MinutosFaltantes > 0.  Puede estar vacío.
    """
    _DIAS_ES = [
        "lunes", "martes", "miércoles", "jueves",
        "viernes", "sábado", "domingo",
    ]
    target_min = work_hours_per_day * 60

    df_lab = df[df["TipoHora"].str.strip() == "Laboral"].copy()
    if df_lab.empty:
        return pd.DataFrame(
            columns=["FechaRegistro", "DiaSemana", "MinutosRegistrados", "MinutosFaltantes"]
        )

    df_lab["_min"] = df_lab["Horas"] * 60 + df_lab["Minutos"]
    totales = df_lab.groupby("FechaRegistro")["_min"].sum().reset_index()
    totales.columns = ["FechaRegistro", "MinutosRegistrados"]

    totales["MinutosFaltantes"] = target_min - totales["MinutosRegistrados"]
    totales = totales[totales["MinutosFaltantes"] > 0].copy()

    def _dia(fecha_str: str) -> str:
        try:
            return _DIAS_ES[datetime.strptime(fecha_str, "%Y-%m-%d").weekday()]
        except ValueError:
            return ""

    totales["DiaSemana"] = totales["FechaRegistro"].apply(_dia)

    return totales[
        ["FechaRegistro", "DiaSemana", "MinutosRegistrados", "MinutosFaltantes"]
    ].reset_index(drop=True)


# ── Función 4: normalizar DataFrame para validator / uploader ─────────────────

def to_bisopi_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normaliza un DataFrame de eventos parseados al esquema exacto que esperan
    validator.validate y uploader.upload.

    Garantías
    ---------
    - Contiene exactamente las columnas de REGISTRO_COLUMNS en ese orden.
    - Horas y Minutos son enteros.
    - HoraInicio queda vacío para toda fila cuyo TipoHora no sea "Adicional".
    - Columnas de texto están stripped y sin NaN.
    - El índice original se preserva (no se hace reset_index) para que el
      caller pueda hacer .update() de vuelta al DataFrame completo.

    Filas con Estado = "Pendiente" → listas para enviar.
    Filas con Estado = "⚠ Incompleto" / "⚠ Sin clasificar" → se incluyen
    en el resultado pero el validador/uploader las ignorará.
    """
    df = df.copy()

    # Asegurar que existen todas las columnas del modelo
    for col in REGISTRO_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    # Horas y Minutos → entero
    df["Horas"]   = pd.to_numeric(df["Horas"],   errors="coerce").fillna(0).astype(int)
    df["Minutos"] = pd.to_numeric(df["Minutos"], errors="coerce").fillna(0).astype(int)

    # HoraInicio solo tiene sentido para horas Adicionales
    df.loc[df["TipoHora"].str.strip() != "Adicional", "HoraInicio"] = ""

    # Columnas de texto: stripear y convertir NaN → ""
    for col in REGISTRO_COLUMNS:
        if col not in ("Horas", "Minutos"):
            df[col] = df[col].fillna("").astype(str).str.strip()

    return df[REGISTRO_COLUMNS]   # preserva el índice original


# ── Función 5: exportar a la plantilla Excel ──────────────────────────────────

def export_to_template(df: pd.DataFrame) -> bytes:
    """
    Escribe el DataFrame en la hoja 'Registro' de la plantilla Excel configurada
    en PLANTILLA_PATH y devuelve el archivo como bytes para descarga.

    Comportamiento
    --------------
    - Localiza la plantilla a través de PLANTILLA_PATH en config.
    - Escribe filas a partir de la fila 3 (fila 1 = título, fila 2 = encabezados).
    - Respeta el orden de columnas definido en la hoja (usa los encabezados para
      posicionar cada campo, no un índice fijo).
    - Sobreescribe solo valores de celdas existentes; filas extra del DataFrame
      se añaden al final sin formato.
    - No modifica la hoja Histórico ni ninguna otra hoja.

    Lanza
    -----
    ValueError — si PLANTILLA_PATH no está configurada o el archivo no existe.
    """
    try:
        from config import PLANTILLA_PATH as _path
    except ImportError:
        _path = None

    if not _path or not os.path.isfile(_path):
        raise ValueError(
            "PLANTILLA_PATH no está configurada o el archivo no existe. "
            "Configura PLANTILLA_PATH en tu archivo .env para usar esta función."
        )

    wb = load_workbook(_path, keep_vba=False, data_only=False)

    if "Registro" not in wb.sheetnames:
        raise ValueError(
            "La plantilla no contiene la hoja 'Registro'. "
            "Verifica que el archivo indicado en PLANTILLA_PATH es la plantilla correcta."
        )

    ws = wb["Registro"]

    # Localizar posiciones de columna desde la fila de encabezados (fila 2)
    col_positions: dict[str, int] = {}
    for cell in ws[_REGISTRO_HEADER_ROW]:
        if cell.value is not None:
            excel_name = str(cell.value).strip()
            col_num = cell.column if isinstance(cell.column, int) else 1
            col_positions[excel_name] = col_num

    # Limpiar filas de datos existentes (solo valores, preservar formato).
    # Se omiten MergedCell — son celdas no-maestras de rangos fusionados y son
    # de solo lectura en openpyxl; el valor lo controla únicamente la celda maestra.
    max_data_row = max(ws.max_row, _REGISTRO_DATA_START)
    for r in range(_REGISTRO_DATA_START, max_data_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell):
                cell.value = None

    # Números de columna que deben resaltarse en amarillo cuando estén vacías
    _highlight_col_nums: set[int] = {
        col_positions[name]
        for name in _HIGHLIGHT_COLS
        if name in col_positions
    }

    # Escribir filas del DataFrame y resaltar celdas vacías obligatorias
    for row_offset, (_, row) in enumerate(df.iterrows()):
        excel_row = _REGISTRO_DATA_START + row_offset
        for internal_col, excel_col_name in _INTERNAL_TO_EXCEL.items():
            col_num = col_positions.get(excel_col_name)
            if col_num is None:
                continue
            cell = ws.cell(row=excel_row, column=col_num)
            if isinstance(cell, MergedCell):
                continue  # celdas fusionadas no-maestras: no se pueden escribir
            val = row.get(internal_col, "")
            cell.value = val if val != "" else None
            # Resaltar en amarillo si es un campo obligatorio y está vacío
            if col_num in _highlight_col_nums and not cell.value:
                cell.fill = _YELLOW_FILL

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
