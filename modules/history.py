# ═══════════════════════════════════════════════════════════
# BISOPI Automator v1.0 — Mayo 2026
# Autor: Edson Leon
# Cargo: Data Insights Lead — Operaciones
# Empresa: Bision Consulting
# Contacto: edson.leon@bisionconsulting.com
# ═══════════════════════════════════════════════════════════

"""
Operaciones sobre el workbook de la plantilla.

Todas las funciones reciben un Workbook ya abierto y lo modifican en memoria.
El guardado (disco o bytes) se delega a modules.file_manager.save_plantilla.

Funciones públicas:
  update_registro   — actualiza Estado y Respuesta API en la hoja Registro
  append_historico  — añade filas al Histórico (con deduplicación)
  clean_registro    — elimina filas con Estado == "✅ Cargado" de la hoja Registro
  generate_download_bytes — atajo que construye un xlsx en memoria desde source_bytes
"""
from __future__ import annotations

import io
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook, Workbook


# ── Constantes ────────────────────────────────────────────────────────────────

_HIST_HEADERS: list[str] = [
    "Proyecto",
    "Grupo Tarea",
    "Tarea",
    "Tipo Hora",
    "Fecha Registro",
    "Horas",
    "Minutos",
    "Hora Inicio",
    "Comentario",
    "Estado",
    "Respuesta API",
    "Fecha Carga",
]

# Nombres internos del DataFrame → nombres Excel (hoja Registro / Histórico)
_RENAME: dict[str, str] = {
    "GrupoTarea":    "Grupo Tarea",
    "TipoHora":      "Tipo Hora",
    "FechaRegistro": "Fecha Registro",
    "HoraInicio":    "Hora Inicio",
    "RespuestaAPI":  "Respuesta API",
}

_COL_ESTADO    = "Estado"
_COL_RESPUESTA = "Respuesta API"

# La hoja Registro tiene: fila 1 = título, fila 2 = encabezados, fila 3+ = datos
_REGISTRO_HEADER_ROW = 2   # 1-based
_REGISTRO_DATA_START = 3   # 1-based

# Índices de los campos de deduplicación dentro de _HIST_HEADERS
_IDX_PROY  = _HIST_HEADERS.index("Proyecto")        # 0
_IDX_TAREA = _HIST_HEADERS.index("Tarea")            # 2
_IDX_FECHA = _HIST_HEADERS.index("Fecha Registro")   # 4
_IDX_TIPO  = _HIST_HEADERS.index("Tipo Hora")        # 3


# ── Helpers internos ──────────────────────────────────────────────────────────

def _find_col(ws, header_row: int, name: str) -> int | None:
    """
    Devuelve el número de columna (1-based) del encabezado `name`
    en la fila `header_row`.  Compatible con openpyxl >= 2.5.
    """
    for cell in ws[header_row]:
        if cell.value is not None and str(cell.value).strip() == name:
            col = cell.column
            if isinstance(col, str):
                from openpyxl.utils import column_index_from_string
                col = column_index_from_string(col)
            return col
    return None


def _find_header_row_hist(ws) -> int:
    """
    Devuelve la fila (1-based) donde col A == "Proyecto" en la hoja Histórico
    (dentro de las primeras 20 filas).  Devuelve 1 si no se encuentra.
    """
    for r in range(1, min((ws.max_row or 0) + 1, 21)):
        if ws.cell(row=r, column=1).value == "Proyecto":
            return r
    return 1


def _ensure_historico(wb) -> object:
    """
    Devuelve la hoja 'Histórico' del workbook.
    Si no existe, la crea con los encabezados _HIST_HEADERS en la fila 1.
    Si existe pero la fila 1 está vacía (ningún valor), escribe los encabezados.
    No añade encabezados si ya hay contenido en la fila 1.
    """
    if "Histórico" in wb.sheetnames:
        ws = wb["Histórico"]
        if not any(cell.value is not None for cell in ws[1]):
            _write_rows(ws, [_HIST_HEADERS], after_row=0)
    else:
        ws = wb.create_sheet("Histórico")
        _write_rows(ws, [_HIST_HEADERS], after_row=0)
    return ws


def _df_to_hist_rows(df: pd.DataFrame, ts: str) -> list[list]:
    """
    Convierte un DataFrame (columnas internas) en filas listas para escritura.
    Usa _HIST_HEADERS como orden de columnas.
    """
    df_h = df.rename(columns=_RENAME).copy()
    df_h["Fecha Carga"] = ts
    rows = []
    for _, row in df_h.iterrows():
        rows.append([row.get(c, "") for c in _HIST_HEADERS])
    return rows


def _last_data_row(ws) -> int:
    """
    Devuelve la última fila (1-based) que contiene un dato real en la hoja
    Histórico, para que _write_rows escriba a continuación.

    Estrategia: localiza la fila de encabezado buscando 'Proyecto' en la
    columna A (dentro de las primeras 20 filas), luego desciende por esa
    misma columna hasta encontrar la primera celda vacía (None o string
    vacío).  Devuelve la fila ANTERIOR a esa celda vacía.

    Así se ignoran títulos decorativos (fila 1), notas de pie y cualquier
    celda con solo formato.
    """
    max_r      = ws.max_row or 0
    header_row = _find_header_row_hist(ws)

    # Primera fila vacía en col A después del encabezado
    for r in range(header_row + 1, max_r + 2):
        val = ws.cell(row=r, column=1).value
        if val is None or (isinstance(val, str) and val.strip() == ""):
            return r - 1   # la fila anterior es la última con dato real

    return max_r


def _write_rows(ws, rows: list[list], after_row: int) -> None:
    """Escribe cada fila de 'rows' en ws empezando en after_row + 1."""
    for i, row_data in enumerate(rows, start=1):
        for j, val in enumerate(row_data, start=1):
            ws.cell(row=after_row + i, column=j).value = val


def _cell_str(val) -> str:
    """Convierte un valor de celda openpyxl a string normalizado para comparación."""
    if val is None:
        return ""
    if hasattr(val, "strftime"):          # datetime / date
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


def _build_existing_keys(ws) -> set[tuple[str, str, str, str]]:
    """
    Lee el Histórico y devuelve el conjunto de claves ya presentes.
    Clave: (Proyecto, Tarea, Fecha Registro, Tipo Hora)
    """
    header_row = _find_header_row_hist(ws)
    c_proy  = _find_col(ws, header_row, "Proyecto")
    c_tarea = _find_col(ws, header_row, "Tarea")
    c_fecha = _find_col(ws, header_row, "Fecha Registro")
    c_tipo  = _find_col(ws, header_row, "Tipo Hora")

    if not all([c_proy, c_tarea, c_fecha, c_tipo]):
        return set()

    keys: set[tuple] = set()
    max_r = ws.max_row or 0
    for r in range(header_row + 1, max_r + 1):
        proy = ws.cell(r, c_proy).value
        # Ignorar filas completamente vacías (solo formato o nota de pie)
        if proy is None or _cell_str(proy) == "":
            continue
        key = (
            _cell_str(ws.cell(r, c_proy).value),
            _cell_str(ws.cell(r, c_tarea).value),
            _cell_str(ws.cell(r, c_fecha).value),
            _cell_str(ws.cell(r, c_tipo).value),
        )
        keys.add(key)
    return keys


# ── Función 1: actualizar Estado y Respuesta API en la hoja Registro ──────────

def update_registro(wb: Workbook, df: pd.DataFrame) -> None:
    """
    Localiza la hoja 'Registro' del workbook y actualiza las columnas
    'Estado' y 'Respuesta API' fila a fila por posición
    (df.iloc[0] ↔ Excel fila 3, df.iloc[1] ↔ Excel fila 4, …).

    No modifica ninguna otra celda ni formato.
    No guarda el workbook — el caller llama a save_plantilla.
    """
    if "Registro" not in wb.sheetnames:
        return

    ws = wb["Registro"]

    col_estado    = _find_col(ws, _REGISTRO_HEADER_ROW, _COL_ESTADO)
    col_respuesta = _find_col(ws, _REGISTRO_HEADER_ROW, _COL_RESPUESTA)

    if col_estado is None or col_respuesta is None:
        return

    for i, (_, row) in enumerate(df.iterrows()):
        excel_row = _REGISTRO_DATA_START + i
        if excel_row > ws.max_row:
            break
        ws.cell(row=excel_row, column=col_estado).value    = str(row.get("Estado",       ""))
        ws.cell(row=excel_row, column=col_respuesta).value = str(row.get("RespuestaAPI", ""))


# ── Función 2: añadir filas al Histórico (con deduplicación) ──────────────────

def append_historico(wb: Workbook, df: pd.DataFrame) -> None:
    """
    Agrega al final de la hoja 'Histórico' todas las filas del DataFrame,
    omitiendo duplicados según la clave (Proyecto, Tarea, FechaRegistro, TipoHora).

    Añade columna 'Fecha Carga' con timestamp actual.
    Si la hoja no existe la crea con encabezados.
    No guarda el workbook — el caller llama a save_plantilla.
    """
    if df.empty:
        return

    ts   = datetime.now().strftime("%d/%m/%Y %H:%M")
    rows = _df_to_hist_rows(df, ts)

    ws           = _ensure_historico(wb)
    existing_keys = _build_existing_keys(ws)

    new_rows: list[list] = []
    for row_data in rows:
        key = (
            _cell_str(row_data[_IDX_PROY]),
            _cell_str(row_data[_IDX_TAREA]),
            _cell_str(row_data[_IDX_FECHA]),
            _cell_str(row_data[_IDX_TIPO]),
        )
        if key in existing_keys:
            print(
                f"Duplicado omitido: {row_data[_IDX_PROY]} / "
                f"{row_data[_IDX_TAREA]} / {row_data[_IDX_FECHA]}"
            )
        else:
            new_rows.append(row_data)
            # NO agregamos key a existing_keys: dos registros del mismo batch
            # con igual (Proyecto, Tarea, Fecha, TipoHora) son legítimamente
            # distintos (diferente duración o comentario).  La protección
            # anti-duplicado aplica solo contra entradas ya persistidas en el sheet.

    if new_rows:
        _write_rows(ws, new_rows, after_row=_last_data_row(ws))


# ── Función 3: eliminar filas cargadas de la hoja Registro ───────────────────

def clean_registro(wb: Workbook) -> int:
    """
    Elimina de la hoja 'Registro' todas las filas cuyo Estado == '✅ Cargado'.
    Las filas se borran de arriba hacia abajo (en orden inverso para no
    desplazar índices).

    Devuelve el número de filas eliminadas.
    No guarda el workbook — el caller llama a save_plantilla.
    """
    if "Registro" not in wb.sheetnames:
        return 0

    ws         = wb["Registro"]
    col_estado = _find_col(ws, _REGISTRO_HEADER_ROW, _COL_ESTADO)

    if col_estado is None:
        return 0

    to_delete: list[int] = []
    for r in range(_REGISTRO_DATA_START, ws.max_row + 1):
        val = ws.cell(row=r, column=col_estado).value
        if val and str(val).strip() == "✅ Cargado":
            to_delete.append(r)

    for r in reversed(to_delete):
        ws.delete_rows(r)

    return len(to_delete)


# ── Función 4: atajo en memoria para descarga ─────────────────────────────────

def generate_download_bytes(
    df: pd.DataFrame,
    source_bytes: bytes | None,
) -> bytes:
    """
    Genera un xlsx en memoria listo para st.download_button.

    Aplica sobre source_bytes (o un workbook mínimo si no hay archivo):
      1. Actualiza Estado y Respuesta API en la hoja Registro.
      2. Añade al Histórico TODAS las filas con Estado == "✅ Cargado".

    Parámetros
    ----------
    df           : DataFrame completo con resultados de la carga.
    source_bytes : Bytes del archivo Excel original. None → workbook mínimo.
    """
    if source_bytes:
        wb = load_workbook(io.BytesIO(source_bytes), keep_vba=False, data_only=False)
        update_registro(wb, df)
    else:
        wb = Workbook()
        default = wb.active
        if default is not None and default.title in ("Sheet", "Sheet1"):
            wb.remove(default)

    df_ok = df[df["Estado"] == "✅ Cargado"]
    if not df_ok.empty:
        append_historico(wb, df_ok)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
