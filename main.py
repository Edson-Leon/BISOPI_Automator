# ═══════════════════════════════════════════════════════════
# BISOPI Automator v1.0 — Mayo 2026
# Autor: Edson Leon
# Cargo: Data Insights Lead — Operaciones
# Empresa: Bision Consulting
# Contacto: edson.leon@bisionconsulting.com
# ═══════════════════════════════════════════════════════════

import io
import os
from datetime import date, timedelta

import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook

from config import (
    BISOPI_API_TOKEN,
    PLANTILLA_PATH,
    has_local_path,
    has_graph_access,
    is_cloud,
    get_email_colaborador,
    AZURE_CLIENT_ID,
    AZURE_TENANT_ID,
)
from modules.loader         import load_from_file, load_from_text, build_summary
from modules.validator      import validate
from modules.uploader       import upload, TokenError
from modules.history        import (
    update_registro,
    append_historico,
    clean_registro,
    generate_download_bytes,
)
from modules.file_manager   import save_plantilla, check_writable
from modules.outlook_parser import parse_ics, calculate_gaps, to_bisopi_df, export_to_template

# Submodo B — Microsoft Graph API (requiere: pip install msal requests)
try:
    from modules.graph_client import (
        authenticate_interactive,
        initiate_device_flow,
        poll_device_flow,
        get_calendar_events,
        events_to_dataframe as graph_events_to_dataframe,
        get_user_email,
        get_user_info,
    )
    _GRAPH_CLIENT_AVAILABLE = True
except ImportError:
    _GRAPH_CLIENT_AVAILABLE = False

# Columnas que el usuario puede editar en la tabla (Estado y RespuestaAPI son auto-generados)
_EDITABLE_COLS = [
    "Proyecto", "GrupoTarea", "Tarea", "TipoHora",
    "FechaRegistro", "Horas", "Minutos",
    "HoraInicio", "Comentario",
]

_LOCAL_MODE = has_local_path()

st.set_page_config(
    page_title="BISOPI Automator",
    page_icon="⚡",
    layout="wide",
)


# ── Estilos globales ──────────────────────────────────────────────────────────
st.markdown("""
<style>
html, body, [class*="css"] { font-family: 'Segoe UI', sans-serif; }

.app-header {
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 60%, #0f3460 100%);
    border-radius: 12px;
    padding: 28px 36px;
    margin-bottom: 24px;
    display: flex;
    align-items: center;
    gap: 20px;
}
.app-header h1 { color:#fff; font-size:2rem; font-weight:700; margin:0; letter-spacing:-0.5px; }
.app-header p  { color:#a0aec0; font-size:0.95rem; margin:6px 0 0 0; }
.badge-user {
    display:inline-flex; align-items:center; gap:6px;
    background:rgba(255,255,255,0.1); border:1px solid rgba(255,255,255,0.15);
    color:#e2e8f0; padding:4px 12px; border-radius:20px;
    font-size:0.82rem; margin-top:10px;
}

.option-card {
    background:#f8fafc; border:1.5px solid #e2e8f0;
    border-radius:10px; padding:22px 24px; height:100%;
    transition:border-color 0.2s;
}
.option-card:hover { border-color:#94a3b8; }
.option-card h4 { color:#1e293b; font-size:0.95rem; font-weight:600; margin:0 0 4px 0; }
.option-card p  { color:#64748b; font-size:0.82rem; margin:0 0 16px 0; }
.option-badge {
    display:inline-block; background:#dbeafe; color:#1d4ed8;
    font-size:0.72rem; font-weight:600; padding:2px 8px;
    border-radius:4px; margin-bottom:14px; letter-spacing:.5px;
    text-transform:uppercase;
}

.coming-soon-card {
    background:#f8fafc; border:1.5px dashed #cbd5e1;
    border-radius:10px; padding:40px; text-align:center;
    max-width:560px; margin:40px auto 0 auto;
}
.coming-soon-card .cs-icon { font-size:2.4rem; margin-bottom:12px; }
.coming-soon-card h3 { color:#334155; font-size:1.1rem; margin:0 0 8px 0; }
.coming-soon-card p  { color:#64748b; font-size:0.88rem; line-height:1.6; margin:0; }
.cs-tag {
    display:inline-block; background:#f1f5f9; color:#64748b;
    border:1px solid #e2e8f0; font-size:0.72rem; font-weight:600;
    padding:3px 10px; border-radius:20px; margin-bottom:16px;
    text-transform:uppercase; letter-spacing:.5px;
}

.stFileUploader label { display:none; }
[data-testid="stDataFrameAddRow"] { display:none !important; }
</style>
""", unsafe_allow_html=True)


# ── Control de sesión — autenticación obligatoria en cloud ────────────────────
if is_cloud():
    if "cloud__user_email" not in st.session_state:
        # ── Página de login (cloud no autenticado) ────────────────────────────
        st.markdown("""
        <div style="background:linear-gradient(135deg,#1a1a2e 0%,#16213e 60%,#0f3460 100%);
                    border-radius:12px;padding:32px 36px;margin-bottom:28px;">
          <div style="color:#ffffff;font-size:1.6rem;font-weight:700;">⚡ BISOPI Automator</div>
          <div style="color:#a0aec0;font-size:0.9rem;margin-top:6px;">
            Para usar la app debes autenticarte con tu cuenta corporativa de Bision.
          </div>
        </div>
        """, unsafe_allow_html=True)

        # ── Selector de método ────────────────────────────────────────────────
        st.radio(
            "Método de autenticación",
            options=[
                "🖥️ Ventana de login de Microsoft",
                "📱 Código de dispositivo (sin ventana emergente)",
            ],
            key="cloud__auth_method",
            horizontal=True,
        )
        _cl_method = st.session_state["cloud__auth_method"]
        st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

        _cl_flow = st.session_state.get("cloud__auth_flow")

        # ── Método: Ventana de login (interactive) ────────────────────────────
        if "Código" not in _cl_method:
            _cl_conn_col, _ = st.columns([3, 9])
            with _cl_conn_col:
                if st.button(
                    "🔗 Conectar con Microsoft",
                    type="primary",
                    key="cloud__btn_connect",
                    use_container_width=True,
                ):
                    with st.spinner("Abriendo el navegador para autenticación…"):
                        try:
                            _cl_token = authenticate_interactive()
                            _cl_user  = get_user_info(_cl_token)
                            st.session_state["cloud__user_email"]    = _cl_user["email"]
                            st.session_state["cloud__user_name"]     = _cl_user["name"]
                            st.session_state["outlook__graph_token"] = _cl_token
                            st.session_state["outlook__graph_email"] = _cl_user["email"]
                            st.rerun()
                        except (RuntimeError, ImportError) as _cl_exc:
                            _cl_err = str(_cl_exc)
                            st.error(_cl_exc)
                            if "código de dispositivo" in _cl_err.lower():
                                st.info(
                                    "💡 Prueba con el método de "
                                    "**código de dispositivo**."
                                )
            st.caption(
                "Se abrirá el navegador para iniciar sesión. Requiere que IT "
                "registre la Redirect URI de esta app en Azure AD."
            )

        # ── Método: Código de dispositivo ────────────────────────────────────
        else:
            if _cl_flow is None:
                # Paso 1 — iniciar flujo
                _cl_conn_col, _ = st.columns([3, 9])
                with _cl_conn_col:
                    if st.button(
                        "🔗 Conectar con Microsoft",
                        type="primary",
                        key="cloud__btn_connect",
                        use_container_width=True,
                    ):
                        try:
                            _cl_flow = initiate_device_flow()
                            st.session_state["cloud__auth_flow"] = _cl_flow
                            st.rerun()
                        except (RuntimeError, ImportError) as _cl_exc:
                            st.error(str(_cl_exc))
                st.caption(
                    "No abre ventanas emergentes — te dará un código para "
                    "introducir en el navegador desde cualquier dispositivo."
                )
            else:
                # Paso 2 — mostrar código y esperar confirmación
                st.info(
                    f"**1.** Abre este enlace en tu navegador: "
                    f"[{_cl_flow.get('verification_uri', 'https://microsoft.com/devicelogin')}]"
                    f"({_cl_flow.get('verification_uri', 'https://microsoft.com/devicelogin')})\n\n"
                    f"**2.** Ingresa el código: `{_cl_flow.get('user_code', '—')}`\n\n"
                    "**3.** El código expira en 15 minutos."
                )
                _cl_done_col, _cl_cancel_col, _ = st.columns([2, 2, 8])
                with _cl_done_col:
                    if st.button(
                        "✅ Ya me autentiqué",
                        type="primary",
                        key="cloud__btn_done",
                        use_container_width=True,
                    ):
                        try:
                            _cl_token = poll_device_flow(_cl_flow)
                            if _cl_token is None:
                                st.info(
                                    "⏳ Aún no detectamos tu autenticación — "
                                    "espera unos segundos e inténtalo de nuevo."
                                )
                            else:
                                _cl_user = get_user_info(_cl_token)
                                st.session_state["cloud__user_email"]    = _cl_user["email"]
                                st.session_state["cloud__user_name"]     = _cl_user["name"]
                                st.session_state["cloud__auth_flow"]     = None
                                st.session_state["outlook__graph_token"] = _cl_token
                                st.session_state["outlook__graph_email"] = _cl_user["email"]
                                st.rerun()
                        except Exception as _cl_exc:
                            st.error(str(_cl_exc))
                with _cl_cancel_col:
                    if st.button(
                        "✖ Cancelar",
                        key="cloud__btn_cancel",
                        use_container_width=True,
                    ):
                        st.session_state["cloud__auth_flow"] = None
                        st.rerun()

        st.stop()

    else:
        # ── Autenticado — email del usuario desde session_state ───────────────
        email_colaborador = st.session_state["cloud__user_email"]
        st.sidebar.success(f"👤 {st.session_state['cloud__user_name']}")
        st.sidebar.caption(email_colaborador)
        if st.sidebar.button("Cerrar sesión", key="cloud__btn_logout"):
            for _k in ["cloud__user_email", "cloud__user_name",
                       "cloud__auth_flow", "outlook__graph_token",
                       "outlook__graph_email"]:
                st.session_state.pop(_k, None)
            st.rerun()

else:
    # ── Local — email desde .env ──────────────────────────────────────────────
    email_colaborador = get_email_colaborador()


# ── Header ────────────────────────────────────────────────────────────────────
user_badge = ""
if email_colaborador:
    user_badge = f'<div class="badge-user">🟢 {email_colaborador}</div>'
else:
    user_badge = (
        '<div class="badge-user" style="background:rgba(239,68,68,0.15);'
        'border-color:rgba(239,68,68,0.3);color:#fca5a5;">'
        '⚠️ Sin configurar — crea tu archivo .env</div>'
    )

st.markdown(f"""
<div class="app-header">
    <div>
        <h1>⚡ BISOPI Automator</h1>
        <p>Carga tu semana de trabajo a BISOPI en un solo paso — sin redigitar, sin recargas.</p>
        {user_badge}
    </div>
</div>
""", unsafe_allow_html=True)


# ── Validación de configuración ───────────────────────────────────────────────
_config_errors: list[str] = []
# En cloud el email viene del login — solo se valida en local
if not is_cloud() and not email_colaborador:
    _config_errors.append("**EMAIL_COLABORADOR** no está definido en el archivo `.env`.")
if not BISOPI_API_TOKEN:
    _config_errors.append("**BISOPI_API_TOKEN** no está definido.")

if _config_errors:
    st.error(
        "⚠️ Configuración incompleta — la aplicación no puede enviar registros sin estas credenciales.\n\n"
        + "\n".join(f"- {e}" for e in _config_errors)
        + "\n\nCrea o edita el archivo **`.env`** en la raíz del proyecto "
          "y reinicia la app con `streamlit run main.py`."
    )
    st.stop()

# Aviso si PLANTILLA_PATH está definida pero el archivo no existe
if PLANTILLA_PATH and not _LOCAL_MODE:
    st.warning(
        f"⚠️ **PLANTILLA_PATH** está configurada (`{PLANTILLA_PATH}`) "
        "pero el archivo no existe en esa ruta. "
        "Se usará el modo de descarga como alternativa."
    )


# ── Tabs de fuente ────────────────────────────────────────────────────────────
tab_archivo, tab_devops, tab_outlook, tab_about = st.tabs(
    ["📂  Archivo plano", "🔷  Azure DevOps", "📅  Agenda Outlook", "ℹ️  Acerca de"]
)


# ── Tab 1: Archivo plano ──────────────────────────────────────────────────────
with tab_archivo:
    st.markdown("<br>", unsafe_allow_html=True)

    # ── Inicializar session state ─────────────────────────────────────────────
    for _key, _val in [
        ("archivo__df_edit",        None),
        ("archivo__src_key",        None),
        ("archivo__source_bytes",   None),
        ("archivo__src_filename",   None),
        ("archivo__upload_result",  None),
        ("archivo__write_ok",       None),   # None=no intentado, True=guardado en disco, False=fallo/descarga
        ("archivo__download_bytes", None),   # bytes para st.download_button (modo descarga)
        ("archivo__clean_confirm",  False),  # estado del diálogo de confirmación "limpiar"
    ]:
        if _key not in st.session_state:
            st.session_state[_key] = _val

    # ─────────────────────────────────────────────────────────────────────────
    # MODO A: ruta local configurada y archivo existente (solo en local)
    # ─────────────────────────────────────────────────────────────────────────
    if not is_cloud() and _LOCAL_MODE:
        col_info, col_btn = st.columns([8, 2])
        with col_info:
            st.success(f"✅ Archivo configurado: `{PLANTILLA_PATH}`")
        with col_btn:
            if st.button("🔄 Recargar archivo", use_container_width=True):
                st.session_state["archivo__src_key"]        = None
                st.session_state["archivo__upload_result"]  = None
                st.session_state["archivo__write_ok"]       = None
                st.session_state["archivo__download_bytes"] = None
                st.session_state["archivo__clean_confirm"]  = False
                st.rerun()

        # Auto-cargar cuando no hay datos o cuando se solicitó recarga
        if st.session_state["archivo__src_key"] != "local_path":
            try:
                with open(PLANTILLA_PATH, "rb") as fh:
                    raw = fh.read()
                st.session_state["archivo__df_edit"]      = load_from_file(PLANTILLA_PATH)
                st.session_state["archivo__source_bytes"] = raw
                st.session_state["archivo__src_key"]      = "local_path"
                st.session_state["archivo__src_filename"] = os.path.basename(PLANTILLA_PATH)
            except ValueError as exc:
                st.error(f"Error al leer el archivo: {exc}")
                st.session_state["archivo__df_edit"] = None
            except Exception as exc:
                st.error(f"No se pudo abrir el archivo: {exc}")
                st.session_state["archivo__df_edit"] = None

    # ─────────────────────────────────────────────────────────────────────────
    # MODO B: modo descarga (file_uploader + copy-paste)
    # ─────────────────────────────────────────────────────────────────────────
    else:
        col_upload, col_or, col_paste = st.columns([10, 1, 10], gap="small")

        with col_upload:
            st.markdown("""
            <div class="option-card">
                <div class="option-badge">Opción A</div>
                <h4>📎 Cargar archivo Excel</h4>
                <p>Sube tu plantilla <code>Plantilla_BISOPI_Automator.xlsx</code> directamente.</p>
            </div>
            """, unsafe_allow_html=True)
            st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
            uploaded_file = st.file_uploader(
                label="xlsx",
                type=["xlsx"],
                label_visibility="collapsed",
                help=(
                    "Solo se procesa la hoja 'Registro'. "
                    "Las columnas deben coincidir con la plantilla oficial."
                ),
            )

        with col_or:
            st.markdown("""
            <div style="display:flex;flex-direction:column;align-items:center;
                        justify-content:center;height:180px;gap:6px;">
                <div style="flex:1;width:1px;background:#e2e8f0;min-height:40px;"></div>
                <div style="color:#94a3b8;font-weight:600;font-size:0.78rem;
                            white-space:nowrap;">ó</div>
                <div style="flex:1;width:1px;background:#e2e8f0;min-height:40px;"></div>
            </div>
            """, unsafe_allow_html=True)

        with col_paste:
            st.markdown("""
            <div class="option-card">
                <div class="option-badge">Opción B</div>
                <h4>📋 Pegar rango de Excel</h4>
                <p>Copia desde la fila de encabezados de la hoja <code>Registro</code>
                   y pega con <kbd>Ctrl+V</kbd>.</p>
            </div>
            """, unsafe_allow_html=True)
            st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
            pasted_text = st.text_area(
                label="paste",
                label_visibility="collapsed",
                height=160,
                placeholder=(
                    "Selecciona desde la fila de encabezados hacia abajo en Excel,\n"
                    "copia (Ctrl+C) y pega aquí (Ctrl+V).\n\n"
                    "Formato esperado: columnas separadas por tabulaciones,\n"
                    "una fila por registro."
                ),
            )
            procesar_texto = st.button(
                "▶ Procesar texto",
                use_container_width=True,
                disabled=not pasted_text.strip(),
            )

        st.markdown(
            "<div style='color:#94a3b8;font-size:0.78rem;text-align:center;margin-top:6px'>"
            "Si cargas archivo y pegas texto al mismo tiempo, el archivo tiene prioridad."
            "</div>",
            unsafe_allow_html=True,
        )

        # Tip para configurar PLANTILLA_PATH
        st.markdown(
            "<div style='color:#64748b;font-size:0.8rem;text-align:center;margin-top:4px'>"
            "💡 Configura <code>PLANTILLA_PATH</code> en tu <code>.env</code> "
            "para evitar cargar y descargar el archivo cada vez."
            "</div>",
            unsafe_allow_html=True,
        )

        # ── Procesamiento de fuente (modo descarga) ───────────────────────────
        if uploaded_file:
            src_key = f"{uploaded_file.name}_{uploaded_file.size}"
            if st.session_state["archivo__src_key"] != src_key:
                try:
                    st.session_state["archivo__df_edit"] = load_from_file(uploaded_file)

                    raw = uploaded_file.getvalue()
                    st.session_state["archivo__src_key"]        = src_key
                    st.session_state["archivo__source_bytes"]   = raw
                    st.session_state["archivo__src_filename"]   = uploaded_file.name
                    st.session_state["archivo__upload_result"]  = None
                    st.session_state["archivo__write_ok"]       = None
                    st.session_state["archivo__download_bytes"] = None
                    st.session_state["archivo__clean_confirm"]  = False
                    st.success(f"✅ **{uploaded_file.name}** cargado correctamente.")
                except ValueError as exc:
                    st.error(f"Error al leer el archivo: {exc}")
                    st.session_state["archivo__df_edit"] = None
            else:
                st.success(f"✅ **{uploaded_file.name}** cargado correctamente.")

        elif procesar_texto and pasted_text.strip():
            try:
                st.session_state["archivo__df_edit"]        = load_from_text(pasted_text)
                st.session_state["archivo__src_key"]        = "text"
                st.session_state["archivo__source_bytes"]   = None
                st.session_state["archivo__src_filename"]   = None
                st.session_state["archivo__upload_result"]  = None
                st.session_state["archivo__write_ok"]       = None
                st.session_state["archivo__download_bytes"] = None
                st.session_state["archivo__clean_confirm"]  = False
                st.success("✅ Texto procesado correctamente.")
            except ValueError as exc:
                st.error(f"Error al procesar el texto: {exc}")
                st.session_state["archivo__df_edit"] = None

    # ── Resultados ────────────────────────────────────────────────────────────
    if st.session_state["archivo__df_edit"] is not None and not st.session_state["archivo__df_edit"].empty:

        df_validado = validate(st.session_state["archivo__df_edit"])

        # ── Resumen de validación ─────────────────────────────────────────────
        n_cargado = int((df_validado["Estado"] == "✅ Cargado").sum())
        n_error   = int((df_validado["Estado"] == "❌ Error").sum())
        n_warn    = int(
            (
                (df_validado["Estado"] == "Pendiente") &
                df_validado["RespuestaAPI"].str.contains("⚠", na=False)
            ).sum()
        )
        n_listo  = int((df_validado["Estado"] == "Pendiente").sum()) - n_warn
        n_subir  = n_listo + n_warn

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("#### Registros cargados")

        badges: list[str] = []
        if n_cargado:
            badges.append(
                f'<span style="background:#DCFCE7;color:#166534;padding:3px 10px;'
                f'border-radius:12px;font-size:0.82rem;font-weight:600">'
                f'✅ {n_cargado} enviado{"s" if n_cargado != 1 else ""}</span>'
            )
        badges.append(
            f'<span style="background:#DCFCE7;color:#166534;padding:3px 10px;'
            f'border-radius:12px;font-size:0.82rem;font-weight:600">'
            f'✅ {n_listo} lista{"s" if n_listo != 1 else ""}</span>'
        )
        badges.append(
            f'<span style="background:#FEF9C3;color:#854D0E;padding:3px 10px;'
            f'border-radius:12px;font-size:0.82rem;font-weight:600">'
            f'⚠ {n_warn} advertencia{"s" if n_warn != 1 else ""}</span>'
        )
        badges.append(
            f'<span style="background:#FEE2E2;color:#991B1B;padding:3px 10px;'
            f'border-radius:12px;font-size:0.82rem;font-weight:600">'
            f'❌ {n_error} con error</span>'
        )
        st.markdown(
            '<div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px">'
            + "".join(badges) + "</div>",
            unsafe_allow_html=True,
        )

        # ── Función de color de filas ─────────────────────────────────────────
        def _style_tabla(row: pd.Series) -> list[str]:
            estado = row["Estado"]
            resp   = str(row.get("RespuestaAPI", ""))
            if estado == "❌ Error":
                color = "background-color: #FEE2E2"
            elif estado == "✅ Cargado":
                color = "background-color: #DCFCE7"
            elif "⚠" in resp:
                color = "background-color: #FEF9C3"
            else:
                color = ""
            return [color] * len(row)

        # ── Vista post-upload (resultados) o tabla editable ───────────────────
        upload_result = st.session_state.get("archivo__upload_result")

        if upload_result is not None:

            # ── Mensaje persistido desde la limpieza (sobrevive st.rerun) ────
            clean_msg = st.session_state.pop("archivo___clean_msg", None)
            if clean_msg:
                st.success(clean_msg)

            # ── Modo resultados ───────────────────────────────────────────────
            df_res = upload_result
            n_ok   = int((df_res["Estado"] == "✅ Cargado").sum())
            n_fail = int((df_res["Estado"] == "❌ Error").sum())

            if n_fail == 0:
                st.success(
                    f"🎉 ¡Todo listo! **{n_ok}** registro{'s' if n_ok != 1 else ''} "
                    f"enviado{'s' if n_ok != 1 else ''} correctamente a BISOPI."
                )
            else:
                st.warning(
                    f"Upload completado con resultados mixtos — "
                    f"**{n_ok}** exitoso{'s' if n_ok != 1 else ''}, "
                    f"**{n_fail}** fallido{'s' if n_fail != 1 else ''}."
                )

            # Mensaje de escritura al archivo
            write_ok = st.session_state.get("archivo__write_ok")
            if _LOCAL_MODE and write_ok is True and n_ok > 0:
                st.success("💾 Archivo actualizado en su ubicación original.")

            # Tabla de resultados (editable: el usuario puede corregir filas ❌ en sitio)
            df_res_display = df_res[
                ["Estado"] + [c for c in df_res.columns if c != "Estado"]
            ]
            edited_res = st.data_editor(
                df_res_display.style.apply(_style_tabla, axis=1),
                use_container_width=True,
                hide_index=True,
                disabled=["Estado", "RespuestaAPI"],
                column_config={
                    "Proyecto":      st.column_config.TextColumn("Proyecto"),
                    "GrupoTarea":    st.column_config.TextColumn("Grupo Tarea"),
                    "Tarea":         st.column_config.TextColumn("Tarea"),
                    "TipoHora":      st.column_config.SelectboxColumn(
                                         "Tipo Hora", options=["Laboral", "Adicional"]),
                    "FechaRegistro": st.column_config.TextColumn("Fecha (YYYY-MM-DD)"),
                    "Horas":         st.column_config.NumberColumn(
                                         "Horas", format="%d h", min_value=0),
                    "Minutos":       st.column_config.SelectboxColumn(
                                         "Minutos", options=[0, 15, 30, 45]),
                    "HoraInicio":    st.column_config.TextColumn("Hora Inicio"),
                    "Comentario":    st.column_config.TextColumn("Comentario"),
                    "Estado":        st.column_config.TextColumn("Estado"),
                    "RespuestaAPI":  st.column_config.TextColumn("Respuesta / Advertencia"),
                },
            )

            # Persistir las ediciones en df_edit para que "Reintentar fallidos" las recoja
            if len(edited_res) == len(df_res):
                _updated = df_res.copy()
                for _col in [c for c in edited_res.columns
                             if c not in ("Estado", "RespuestaAPI")]:
                    if _col in _updated.columns:
                        _updated[_col] = edited_res[_col].values
                st.session_state["archivo__df_edit"] = _updated

            st.markdown("<br>", unsafe_allow_html=True)

            # ── Botones de acción ─────────────────────────────────────────────
            action_cols = st.columns([2, 2, 2, 4])

            # Botón: Reintentar fallidos
            with action_cols[0]:
                if n_fail > 0 and st.button(
                    f"🔄 Reintentar fallidos ({n_fail})",
                    use_container_width=True,
                ):
                    df_retry = st.session_state.get("archivo__df_edit", df_res).copy()
                    mask_err = df_retry["Estado"] == "❌ Error"
                    df_retry.loc[mask_err, "Estado"]       = "Pendiente"
                    df_retry.loc[mask_err, "RespuestaAPI"] = ""
                    st.session_state["archivo__df_edit"]        = df_retry
                    st.session_state["archivo__upload_result"]  = None
                    st.session_state["archivo__write_ok"]       = None
                    st.session_state["archivo__download_bytes"] = None
                    st.session_state["archivo__clean_confirm"]  = False
                    st.rerun()

            # Botón: Descargar (solo modo descarga)
            with action_cols[1]:
                if not _LOCAL_MODE:
                    dl_bytes = st.session_state.get("archivo__download_bytes")
                    if dl_bytes:
                        src_name = st.session_state.get("archivo__src_filename") or "BISOPI"
                        dl_name  = src_name.replace(".xlsx", "") + "_actualizado.xlsx"
                        st.download_button(
                            label="⬇ Descargar archivo actualizado",
                            data=dl_bytes,
                            file_name=dl_name,
                            mime=(
                                "application/vnd.openxmlformats-"
                                "officedocument.spreadsheetml.sheet"
                            ),
                            use_container_width=True,
                            help=(
                                "Incluye los estados actualizados en la hoja Registro "
                                "y los registros exitosos en la hoja Histórico."
                            ),
                        )

            # Botón: Limpiar semana cargada (solo cuando hay filas ✅)
            with action_cols[2]:
                if n_ok > 0:
                    if not st.session_state.get("archivo__clean_confirm"):
                        if st.button(
                            "🧹 Limpiar semana cargada",
                            use_container_width=True,
                            help="Elimina del Registro las filas ya enviadas a BISOPI.",
                        ):
                            st.session_state["archivo__clean_confirm"] = True
                            st.rerun()
                    else:
                        pass   # el diálogo se renderiza abajo

            # ── Diálogo de confirmación limpieza ──────────────────────────────
            if st.session_state.get("archivo__clean_confirm") and n_ok > 0:
                st.warning(
                    f"⚠️ ¿Confirmas eliminar **{n_ok}** "
                    f"fila{'s' if n_ok != 1 else ''} cargada{'s' if n_ok != 1 else ''} "
                    "de la hoja Registro?"
                )
                conf_col, cancel_col, _ = st.columns([2, 2, 6])

                with conf_col:
                    if st.button("✅ Confirmar limpieza", type="primary", use_container_width=True):

                        # ── Pre-check: detectar archivo bloqueado ANTES de modificar ──
                        if _LOCAL_MODE and not check_writable(PLANTILLA_PATH):
                            st.warning(
                                "⚠️ **El archivo está abierto en Excel u otro programa.** "
                                "Ciérralo y vuelve a intentarlo."
                            )
                            # No hacer st.rerun() — dejar el warning visible
                        else:
                            # Abrir workbook desde la fuente más actualizada
                            if _LOCAL_MODE:
                                wb_clean = load_workbook(
                                    PLANTILLA_PATH, keep_vba=False, data_only=False
                                )
                            else:
                                dl_bytes = st.session_state.get("archivo__download_bytes")
                                if dl_bytes:
                                    wb_clean = load_workbook(
                                        io.BytesIO(dl_bytes), keep_vba=False, data_only=False
                                    )
                                else:
                                    wb_clean = load_workbook(
                                        io.BytesIO(st.session_state["archivo__source_bytes"]),
                                        keep_vba=False, data_only=False,
                                    )

                            n_deleted = clean_registro(wb_clean)

                            if _LOCAL_MODE:
                                clean_ok, _ = save_plantilla(wb_clean, PLANTILLA_PATH)
                            else:
                                clean_ok, new_bytes = save_plantilla(wb_clean, None)
                                if new_bytes:
                                    st.session_state["archivo__download_bytes"] = new_bytes
                                    clean_ok = True

                            st.session_state["archivo__clean_confirm"] = False

                            if clean_ok:
                                # Refrescar source_bytes y df_edit para que el estado
                                # interno refleje exactamente el archivo limpiado.
                                # Sin esto, el próximo guardado parte de bytes viejos
                                # y restaura las filas que se acaban de eliminar.
                                if _LOCAL_MODE:
                                    with open(PLANTILLA_PATH, "rb") as _fh:
                                        fresh_bytes = _fh.read()
                                    st.session_state["archivo__source_bytes"] = fresh_bytes
                                    try:
                                        st.session_state["archivo__df_edit"] = load_from_file(
                                            PLANTILLA_PATH
                                        )
                                    except Exception:
                                        pass  # si falla, df_edit queda con el estado anterior

                                st.session_state["archivo___clean_msg"] = (
                                    f"✅ {n_deleted} fila{'s' if n_deleted != 1 else ''} "
                                    f"cargada{'s' if n_deleted != 1 else ''} eliminada{'s' if n_deleted != 1 else ''} "
                                    "del Registro."
                                )
                                st.rerun()
                            # Si clean_ok es False: save_plantilla ya mostró st.warning,
                            # NO llamamos st.rerun() para que el aviso permanezca visible.

                with cancel_col:
                    if st.button("✖ Cancelar", use_container_width=True):
                        st.session_state["archivo__clean_confirm"] = False
                        st.rerun()

        else:
            # ── Modo edición: tabla editable ──────────────────────────────────
            df_display = df_validado[
                ["Estado"] + [c for c in df_validado.columns if c != "Estado"]
            ]
            edited_df = st.data_editor(
                df_display.style.apply(_style_tabla, axis=1),
                use_container_width=True,
                hide_index=True,
                num_rows="dynamic",
                disabled=["Estado", "RespuestaAPI"],
                column_config={
                    "Proyecto":      st.column_config.TextColumn("Proyecto"),
                    "GrupoTarea":    st.column_config.TextColumn("Grupo Tarea"),
                    "Tarea":         st.column_config.TextColumn("Tarea"),
                    "TipoHora":      st.column_config.SelectboxColumn(
                                         "Tipo Hora", options=["Laboral", "Adicional"]),
                    "FechaRegistro": st.column_config.TextColumn("Fecha (YYYY-MM-DD)"),
                    "Horas":         st.column_config.NumberColumn(
                                         "Horas", format="%d h", min_value=0),
                    "Minutos":       st.column_config.SelectboxColumn(
                                         "Minutos", options=[0, 15, 30, 45]),
                    "HoraInicio":    st.column_config.TextColumn("Hora Inicio"),
                    "Comentario":    st.column_config.TextColumn("Comentario"),
                    "Estado":        st.column_config.TextColumn("Estado"),
                    "RespuestaAPI":  st.column_config.TextColumn("Respuesta / Advertencia"),
                },
            )

            # ── Persistir ediciones ───────────────────────────────────────────
            new_edit = edited_df[_EDITABLE_COLS].copy()
            old_edit = st.session_state["archivo__df_edit"]
            old_cols = [c for c in _EDITABLE_COLS if c in old_edit.columns]

            # Estado y RespuestaAPI se leen directamente de edited_df: Streamlit
            # siempre devuelve las columnas "disabled" con sus valores actuales,
            # tanto en ediciones normales como al eliminar filas.
            # Esto evita que las filas ✅ Cargado reviertan a Pendiente al borrar otra.
            if "Estado" in edited_df.columns:
                new_edit["Estado"] = edited_df["Estado"].values
            if "RespuestaAPI" in edited_df.columns:
                new_edit["RespuestaAPI"] = edited_df["RespuestaAPI"].values

            # Si la longitud cambió (fila eliminada) siempre hacer rerun;
            # si es igual, comparar solo columnas editables para detectar cambios.
            if len(new_edit) != len(old_edit):
                changed = True
            else:
                changed = (
                    new_edit[old_cols].astype(str).values.tolist()
                    != old_edit[old_cols].astype(str).values.tolist()
                )

            st.session_state["archivo__df_edit"] = new_edit.reset_index(drop=True)

            if edited_df is not None:
                # Resetear índices para evitar desalineación tras eliminación de filas
                edited_df = edited_df.reset_index(drop=True)
                prev_df = st.session_state["archivo__df_edit"].reset_index(drop=True)

                # Preservar estados ✅ Cargado solo para filas que aún existen en edited_df
                common_idx = edited_df.index.intersection(prev_df.index)
                mask_ok = prev_df.loc[common_idx, "Estado"] == "✅ Cargado"
                ok_idx = common_idx[mask_ok]
                edited_df.loc[ok_idx, "Estado"] = "✅ Cargado"
                edited_df.loc[ok_idx, "RespuestaAPI"] = prev_df.loc[ok_idx, "RespuestaAPI"]
                st.session_state["archivo__df_edit"] = edited_df

            if changed:
                st.rerun()

            # ── Botón Subir a BISOPI ──────────────────────────────────────────
            st.markdown("<br>", unsafe_allow_html=True)

            if n_subir > 0:
                btn_label = (
                    f"🚀 Subir a BISOPI  "
                    f"({n_subir} registro{'s' if n_subir != 1 else ''})"
                )
                if st.button(btn_label, type="primary"):

                    prog_bar  = st.progress(0.0)
                    prog_text = st.empty()

                    def _on_progress(current: int, total: int, label: str) -> None:
                        prog_bar.progress(current / total if total > 0 else 0.0)
                        prog_text.caption(f"Enviando {current} de {total}: *{label}*")

                    try:
                        # ── Pre-check archivo bloqueado (modo local) ──────────
                        if _LOCAL_MODE and not check_writable(PLANTILLA_PATH):
                            st.warning(
                                "⚠️ **El archivo está abierto en Excel u otro programa.** "
                                "Ciérralo y vuelve a intentarlo."
                            )
                            st.stop()

                        # Solo las filas Pendiente se envían a la API
                        df_to_upload = st.session_state["archivo__df_edit"][
                            st.session_state["archivo__df_edit"]["Estado"] == "Pendiente"
                        ].copy()

                        df_result = upload(
                            df_to_upload,
                            email_colaborador,
                            BISOPI_API_TOKEN,
                            _on_progress,
                        )
                        prog_bar.progress(1.0)
                        prog_text.caption("✅ Proceso de envío completado.")

                        # Merge los resultados de vuelta al DataFrame completo por índice
                        st.session_state["archivo__df_edit"].update(df_result)

                        # Filas recién exitosas en este batch (no duplicar en Histórico).
                        # Se filtran por índice de df_to_upload para incluir solo las que
                        # cambiaron en este envío, evitando duplicar ✅ de batches anteriores.
                        df_new_ok = st.session_state["archivo__df_edit"].loc[
                            df_to_upload.index[
                                st.session_state["archivo__df_edit"].loc[df_to_upload.index, "Estado"] == "✅ Cargado"
                            ]
                        ]

                        # DataFrame completo (✅ previos + resultados nuevos) para
                        # escribir al Excel y mostrar en la vista de resultados
                        df_full = st.session_state["archivo__df_edit"].copy()

                        # ── Construir workbook con los resultados aplicados ────
                        source_bytes = st.session_state.get("archivo__source_bytes")
                        if source_bytes:
                            wb_result = load_workbook(
                                io.BytesIO(source_bytes), keep_vba=False, data_only=False
                            )
                        else:
                            wb_result = Workbook()
                            _default = wb_result.active
                            if _default and _default.title in ("Sheet", "Sheet1"):
                                wb_result.remove(_default)

                        update_registro(wb_result, df_full)
                        append_historico(wb_result, df_new_ok)

                        # ── Guardar: disco (local) o bytes (descarga) ─────────
                        effective_path = PLANTILLA_PATH if _LOCAL_MODE else None
                        write_ok, dl_bytes = save_plantilla(wb_result, effective_path)

                        # Refrescar source_bytes desde disco para que el próximo
                        # save parta del estado actual, no del archivo original.
                        if write_ok and _LOCAL_MODE:
                            with open(PLANTILLA_PATH, "rb") as _fh:
                                st.session_state["archivo__source_bytes"] = _fh.read()

                        # En modo local: si el guardado falló (save_plantilla ya
                        # mostró st.warning), igualmente persistimos upload_result
                        # para que el usuario vea los resultados del envío.
                        st.session_state["archivo__upload_result"]  = df_full
                        st.session_state["archivo__write_ok"]       = write_ok
                        st.session_state["archivo__download_bytes"] = dl_bytes
                        st.rerun()

                    except TokenError:
                        st.error(
                            "❌ **Token inválido (401 Unauthorized).** "
                            "Verifica el valor de **BISOPI_API_TOKEN** "
                            "en tu archivo `.env` y reinicia la aplicación."
                        )
                        st.stop()

            else:
                st.warning(
                    "No hay registros listos para enviar. "
                    "Corrige los errores marcados en rojo antes de continuar."
                )

        # ── Panel de resumen por día ──────────────────────────────────────────
        smr = build_summary(st.session_state["archivo__df_edit"].copy())

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("#### Resumen de la semana")

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Registros pendientes", smr["total_registros"])
        c2.metric("Horas laborales",       smr["horas_laborales_fmt"])
        c3.metric("Horas adicionales",     smr["horas_adicionales_fmt"])
        c4.metric("Días con registros",    smr["dias_distintos"])

        st.markdown("<br>", unsafe_allow_html=True)

        def _style_resumen(df: pd.DataFrame) -> pd.DataFrame:
            styles = pd.DataFrame("", index=df.index, columns=df.columns)
            for i, mins in enumerate(smr["laboral_mins_por_dia"]):
                if mins > 480:
                    styles.loc[i, "Horas Laborales"] = (
                        "background-color: #FEF9C3; color: #854D0E; font-weight: 600"
                    )
            styles.iloc[-1] = "font-weight: bold; background-color: #F1F5F9"
            return styles

        st.dataframe(
            smr["resumen_df"].style.apply(_style_resumen, axis=None),
            use_container_width=True,
            hide_index=True,
        )

    elif st.session_state.get("archivo__df_edit") is not None and st.session_state["archivo__df_edit"].empty:
        st.warning("El archivo no contiene filas de datos.")


# ── Tab 2: Azure DevOps ───────────────────────────────────────────────────────
with tab_devops:
    st.markdown("""
    <div class="coming-soon-card">
        <div class="cs-icon">🔷</div>
        <div class="cs-tag">V2 — Próximamente</div>
        <h3>Integración con Azure DevOps</h3>
        <p>
            Importará work items y registros de tiempo directamente desde tus
            proyectos de DevOps, mapeando automáticamente los campos hacia la
            estructura de BISOPI.<br><br>
            Requiere token de DevOps y definición del mapeo de tareas.
        </p>
    </div>
    """, unsafe_allow_html=True)


# ── Tab 3: Agenda Outlook ─────────────────────────────────────────────────────
with tab_outlook:
    st.markdown("<br>", unsafe_allow_html=True)

    # ── Inicializar session state ─────────────────────────────────────────────
    for _key, _val in [
        ("outlook__df_edit",        None),
        ("outlook__src_key",        None),
        ("outlook__ics_bytes",      None),
        ("outlook__ics_name",       None),
        ("outlook__upload_result",  None),
        ("outlook__write_ok",       None),
        ("outlook__download_bytes", None),
        ("outlook__upload_confirm", False),   # False | True (dialog) | "go" (ejecutar)
        ("outlook__graph_token",    None),    # access token MSAL o None si no autenticado
        ("outlook__graph_email",    ""),      # email extraído del JWT
        ("outlook__active_submodo", None),    # submodo que pobló outlook__df_edit
        ("outlook__auth_method",    "🖥️ Ventana de login de Microsoft"),  # método seleccionado
        ("outlook__auth_flow",      None),    # dict de Device Code Flow en curso
    ]:
        if _key not in st.session_state:
            st.session_state[_key] = _val

    # ── Selector de submodo ───────────────────────────────────────────────────
    _ol_submodo = st.radio(
        "Submodo de importación",
        options=["📎  Archivo ICS", "🔗  Microsoft Graph API"],
        horizontal=True,
        label_visibility="collapsed",
    )

    # ── Detectar cambio de submodo y limpiar datos del anterior ─────────────
    _ol_prev_submodo = st.session_state.get("outlook__active_submodo")
    if _ol_prev_submodo is not None and _ol_prev_submodo != _ol_submodo:
        st.session_state["outlook__df_edit"]        = None
        st.session_state["outlook__src_key"]        = None
        st.session_state["outlook__ics_bytes"]      = None
        st.session_state["outlook__ics_name"]       = None
        st.session_state["outlook__upload_result"]  = None
        st.session_state["outlook__write_ok"]       = None
        st.session_state["outlook__download_bytes"] = None
        st.session_state["outlook__active_submodo"] = None

    # ── Selector de semana (compartido por ambos submodos) ────────────────────
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    _ol_today  = date.today()
    _ol_monday = _ol_today - timedelta(days=_ol_today.weekday())
    _ol_sunday = _ol_monday + timedelta(days=6)

    _col_ws, _col_we, _col_spacer = st.columns([3, 3, 6])
    with _col_ws:
        _week_start = st.date_input(
            "Inicio de semana",
            value=_ol_monday,
            format="DD/MM/YYYY",
            key="outlook__week_start",
            help="Primer día de la semana a importar (normalmente lunes).",
        )
    with _col_we:
        _week_end = st.date_input(
            "Fin de semana",
            value=_ol_sunday,
            format="DD/MM/YYYY",
            key="outlook__week_end",
            help="Último día de la semana a importar (normalmente domingo).",
        )

    if _week_end < _week_start:
        st.warning("⚠️ La fecha de fin debe ser igual o posterior a la de inicio.")

    # ── Submodo B: Microsoft Graph API ───────────────────────────────────────
    if "Graph" in _ol_submodo:
        if not _GRAPH_CLIENT_AVAILABLE:
            st.error(
                "❌ Las librerías `msal` y/o `requests` no están instaladas. "
                "Ejecuta `pip install msal requests` y reinicia la aplicación."
            )

        elif not has_graph_access():
            st.markdown("""
            <div class="option-card" style="max-width:560px;margin:16px 0;">
              <div style="font-size:1.4rem;margin-bottom:10px;">🔒</div>
              <h4>Credenciales de Azure AD no configuradas</h4>
              <p style="color:#475569;font-size:0.88rem;margin:8px 0 0 0;">
                Para activar esta función, solicita a IT el <strong>Client ID</strong>
                y el <strong>Tenant ID</strong> de la app registrada en Azure AD,
                y agrégalos a tu archivo <code>.env</code>:<br><br>
                <code>AZURE_CLIENT_ID=&lt;valor entregado por IT&gt;</code><br>
                <code>AZURE_TENANT_ID=&lt;valor entregado por IT&gt;</code><br><br>
                Reinicia la aplicación tras guardar el archivo.
              </p>
            </div>
            """, unsafe_allow_html=True)
            st.button(
                "🔒 Conectar con Microsoft",
                disabled=True,
                key="outlook__btn_graph_disabled",
                help="Configura AZURE_CLIENT_ID y AZURE_TENANT_ID en .env para habilitar.",
            )

        else:
            # ── Conectado ─────────────────────────────────────────────────────
            if st.session_state.get("outlook__graph_token"):
                _g_email = st.session_state.get("outlook__graph_email", "")
                _g_info_col, _g_disc_col, _ = st.columns([4, 2, 6])
                with _g_info_col:
                    st.markdown(
                        f'<div class="badge-user" style="display:inline-flex;'
                        f'margin-top:6px;">✅&nbsp;'
                        f'{_g_email or "Conectado a Microsoft 365"}</div>',
                        unsafe_allow_html=True,
                    )
                with _g_disc_col:
                    if st.button(
                        "🔓 Cerrar sesión",
                        key="outlook__btn_graph_disconnect",
                        use_container_width=True,
                    ):
                        st.session_state["outlook__graph_token"]    = None
                        st.session_state["outlook__graph_email"]    = ""
                        st.session_state["outlook__df_edit"]        = None
                        st.session_state["outlook__src_key"]        = None
                        st.session_state["outlook__active_submodo"] = None
                        st.rerun()

                st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

                # Clave que identifica el rango actualmente cargado
                _g_src_key      = f"graph_{_week_start}_{_week_end}"
                _g_range_loaded = (
                    st.session_state.get("outlook__src_key") == _g_src_key
                    and st.session_state.get("outlook__df_edit") is not None
                )

                _g_fetch_col, _ = st.columns([3, 9])
                with _g_fetch_col:
                    _g_btn_label = (
                        "🔄 Actualizar eventos"
                        if _g_range_loaded
                        else "📅 Obtener eventos del calendario"
                    )
                    if st.button(
                        _g_btn_label,
                        type="secondary" if _g_range_loaded else "primary",
                        key="outlook__btn_graph_fetch",
                        use_container_width=True,
                    ):
                        with st.spinner("Obteniendo eventos desde Microsoft 365…"):
                            try:
                                _g_events = get_calendar_events(
                                    st.session_state["outlook__graph_token"],
                                    _week_start,
                                    _week_end,
                                )
                                _g_df = graph_events_to_dataframe(
                                    _g_events, _week_start, _week_end
                                )
                                st.session_state["outlook__df_edit"]        = _g_df
                                st.session_state["outlook__src_key"]        = _g_src_key
                                st.session_state["outlook__active_submodo"] = _ol_submodo
                                st.session_state["outlook__upload_result"]  = None
                                st.session_state["outlook__write_ok"]       = None
                                st.session_state["outlook__download_bytes"] = None
                                st.rerun()
                            except RuntimeError as _g_exc:
                                _g_err_msg = str(_g_exc)
                                if "401" in _g_err_msg:
                                    # Token expirado — forzar re-autenticación
                                    st.session_state["outlook__graph_token"] = None
                                    st.session_state["outlook__graph_email"] = ""
                                    st.warning(
                                        "⚠️ La sesión de Microsoft expiró. "
                                        "Haz clic en **Conectar con Microsoft** para renovarla."
                                    )
                                else:
                                    st.error(_g_err_msg)

            # ── No conectado ───────────────────────────────────────────────────
            else:
                # ── Selector de método de autenticación ───────────────────────
                if is_cloud():
                    st.info(
                        "☁️ En la versión Cloud solo está disponible el método "
                        "de código de dispositivo."
                    )
                    _g_auth_method = "📱 Código de dispositivo (sin ventana emergente)"
                else:
                    st.radio(
                        "Método de autenticación",
                        options=[
                            "🖥️ Ventana de login de Microsoft",
                            "📱 Código de dispositivo (sin ventana emergente)",
                        ],
                        key="outlook__auth_method",
                        horizontal=True,
                    )
                    _g_auth_method = st.session_state["outlook__auth_method"]
                st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

                # ── Método: Ventana de login (interactive) ────────────────────
                if "Código" not in _g_auth_method:
                    _g_conn_col, _ = st.columns([3, 9])
                    with _g_conn_col:
                        if st.button(
                            "🔗 Conectar con Microsoft",
                            type="primary",
                            key="outlook__btn_graph_connect",
                            use_container_width=True,
                        ):
                            with st.spinner(
                                "Abriendo el navegador para autenticación…"
                            ):
                                try:
                                    _g_token = authenticate_interactive()
                                    _g_email = get_user_email(_g_token)
                                    st.session_state["outlook__graph_token"] = _g_token
                                    st.session_state["outlook__graph_email"] = _g_email
                                    st.rerun()
                                except (RuntimeError, ImportError) as _g_exc:
                                    st.error(str(_g_exc))
                                    if "código de dispositivo" in str(_g_exc).lower():
                                        st.info(
                                            "💡 Prueba con el método de "
                                            "**código de dispositivo**"
                                        )
                    st.caption(
                        "Se abrirá el navegador para iniciar sesión con tu cuenta de "
                        "Microsoft 365. El token se guarda en caché local y no se "
                        "transmite fuera de tu equipo."
                    )

                # ── Método: Código de dispositivo (device flow) ───────────────
                else:
                    _g_flow = st.session_state.get("outlook__auth_flow")

                    if _g_flow is None:
                        # Paso 1 — botón para iniciar el flujo
                        _g_conn_col, _ = st.columns([3, 9])
                        with _g_conn_col:
                            if st.button(
                                "🔗 Conectar con Microsoft",
                                type="primary",
                                key="outlook__btn_graph_connect",
                                use_container_width=True,
                            ):
                                try:
                                    _g_flow = initiate_device_flow()
                                    st.session_state["outlook__auth_flow"] = _g_flow
                                    st.rerun()
                                except (RuntimeError, ImportError) as _g_exc:
                                    st.error(str(_g_exc))
                        st.caption(
                            "No abre ventanas emergentes — te dará un código para "
                            "introducir en el navegador desde cualquier dispositivo."
                        )

                    else:
                        # Paso 2 — mostrar instrucciones y esperar confirmación
                        st.info(
                            f"**1.** Abre este enlace en tu navegador: "
                            f"[{_g_flow.get('verification_uri', 'https://microsoft.com/devicelogin')}]"
                            f"({_g_flow.get('verification_uri', 'https://microsoft.com/devicelogin')})\n\n"
                            f"**2.** Ingresa el código: `{_g_flow.get('user_code', '—')}`\n\n"
                            f"**3.** El código expira en 15 minutos."
                        )
                        _g_done_col, _g_cancel_col, _ = st.columns([2, 2, 8])
                        with _g_done_col:
                            if st.button(
                                "✅ Ya me autentiqué",
                                type="primary",
                                key="outlook__btn_graph_device_done",
                                use_container_width=True,
                            ):
                                try:
                                    _g_token = poll_device_flow(_g_flow)
                                    if _g_token is None:
                                        st.info(
                                            "⏳ Aún no detectamos tu autenticación — "
                                            "espera unos segundos e inténtalo de nuevo."
                                        )
                                    else:
                                        _g_email = get_user_email(_g_token)
                                        st.session_state["outlook__graph_token"] = _g_token
                                        st.session_state["outlook__graph_email"] = _g_email
                                        st.session_state["outlook__auth_flow"]   = None
                                        st.rerun()
                                except Exception as _g_exc:
                                    st.error(str(_g_exc))
                        with _g_cancel_col:
                            if st.button(
                                "✖ Cancelar",
                                key="outlook__btn_graph_device_cancel",
                                use_container_width=True,
                            ):
                                st.session_state["outlook__auth_flow"] = None
                                st.rerun()

    # ── Submodo A: Archivo ICS ────────────────────────────────────────────────
    else:
        # ── Instrucciones ─────────────────────────────────────────────────────
        with st.expander("ℹ️  ¿Cómo exportar mi calendario desde Outlook?"):
            st.markdown("""
**Outlook Web:**
1. Ve a Outlook e inicia sesión.
2. Abre la vista de **Calendario** y haz clic en el icono de **Configuración (⚙️)** en la esquina superior derecha.
3. Selecciona **Calendarios → Calendarios compartidos**.
4. En "Publicar un calendario", elige el calendario, selecciona los permisos y haz clic en **Publicar**.
5. Aparecerá un enlace `.ics` — cópialo y pégalo en la barra de direcciones del navegador para descargarlo.

**Outlook Escritorio (Windows):**
1. Abre Outlook y ve a **Archivo → Guardar calendario**.
2. Elige el nombre del archivo y haz clic en **Más opciones** para seleccionar el rango de fechas.
3. Haz clic en **Aceptar** y luego en **Guardar**.

---

**Formato requerido en los eventos:**

El **título** del evento debe comenzar con `[BISOPI]`:
```
[BISOPI] Revisión de sprint con el equipo
```

La **descripción** puede incluir los siguientes campos (uno por línea):
```
Proyecto: Nombre del Proyecto
Grupo Tarea: Nombre del Grupo
Tarea: Descripción detallada de la actividad
Tipo Hora: Laboral
Comentario: Notas adicionales
```

- Si **Tipo Hora** es `Adicional`, el campo **Comentario** es obligatorio y
  se usará la hora de inicio del evento como **Hora Inicio**.
- Los campos no incluidos en la descripción quedan en blanco y pueden
  completarse en la tabla editable antes de enviar.
- Eventos de día completo se convierten automáticamente en registros de **8 horas Laborales**.
- La duración se redondea hacia abajo al múltiplo de **15 minutos** más cercano.
            """)

        # ── Cargador de archivo ICS ───────────────────────────────────────────
        st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
        _ics_file = st.file_uploader(
            "Cargar archivo ICS",
            type=["ics"],
            label_visibility="collapsed",
            help="Exporta tu calendario desde Outlook como .ics y cárgalo aquí.",
        )

        # ── Procesar ICS cuando hay archivo nuevo o semana cambió ─────────────
        if _ics_file is not None:
            _ol_src_key = f"{_ics_file.name}_{_ics_file.size}_{_week_start}_{_week_end}"
            if st.session_state["outlook__src_key"] != _ol_src_key:
                try:
                    _ol_raw = _ics_file.getvalue()
                    _ol_df_parsed = parse_ics(io.BytesIO(_ol_raw), _week_start, _week_end)
                    st.session_state["outlook__df_edit"]        = _ol_df_parsed
                    st.session_state["outlook__ics_bytes"]      = _ol_raw
                    st.session_state["outlook__ics_name"]       = _ics_file.name
                    st.session_state["outlook__src_key"]        = _ol_src_key
                    st.session_state["outlook__active_submodo"] = _ol_submodo
                    st.session_state["outlook__upload_result"]  = None
                    st.session_state["outlook__write_ok"]       = None
                    st.session_state["outlook__download_bytes"] = None
                except ImportError as _exc:
                    st.error(str(_exc))
                    st.session_state["outlook__df_edit"] = None
                except ValueError as _exc:
                    st.error(f"Error al parsear el archivo ICS: {_exc}")
                    st.session_state["outlook__df_edit"] = None

        elif st.session_state.get("outlook__ics_bytes"):
            # Archivo ya procesado: re-parsear si la semana cambió
            _ol_src_key = (
                f"{st.session_state['outlook__ics_name']}_stored_"
                f"{_week_start}_{_week_end}"
            )
            if st.session_state["outlook__src_key"] != _ol_src_key:
                try:
                    _ol_df_parsed = parse_ics(
                        io.BytesIO(st.session_state["outlook__ics_bytes"]),
                        _week_start,
                        _week_end,
                    )
                    st.session_state["outlook__df_edit"]        = _ol_df_parsed
                    st.session_state["outlook__src_key"]        = _ol_src_key
                    st.session_state["outlook__active_submodo"] = _ol_submodo
                    st.session_state["outlook__upload_result"]  = None
                    st.session_state["outlook__write_ok"]       = None
                except Exception as _exc:
                    st.error(f"Error al re-parsear el archivo ICS: {_exc}")

    # ── Resultados (compartido por Submodo A y B) ─────────────────────────────
    def _render_ol_results() -> None:
        """Muestra tabla editable, botones de acción y resumen de la semana."""
        _ol_current_df = st.session_state.get("outlook__df_edit")

        if _ol_current_df is not None:
            if _ol_current_df.empty:
                st.info(
                    "No se encontraron eventos en el rango de fechas seleccionado. "
                    "Verifica que el rango es correcto o que el calendario tiene eventos."
                )
            else:
                # ── Validar ───────────────────────────────────────────────────
                _ol_validado = validate(_ol_current_df)

                _ol_n_ok    = int((_ol_validado["Estado"] == "✅ Cargado").sum())
                _ol_n_err   = int((_ol_validado["Estado"] == "❌ Error").sum())
                _ol_n_warn  = int(
                    ((_ol_validado["Estado"] == "Pendiente") &
                     _ol_validado["RespuestaAPI"].str.contains("⚠", na=False)).sum()
                )
                _ol_n_listo = int((_ol_validado["Estado"] == "Pendiente").sum()) - _ol_n_warn
                _ol_n_subir = _ol_n_listo + _ol_n_warn

                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown("#### Eventos importados")

                _ol_badges: list[str] = []
                if _ol_n_ok:
                    _ol_badges.append(
                        f'<span style="background:#DCFCE7;color:#166534;padding:3px 10px;'
                        f'border-radius:12px;font-size:0.82rem;font-weight:600">'
                        f'✅ {_ol_n_ok} enviado{"s" if _ol_n_ok != 1 else ""}</span>'
                    )
                _ol_badges.append(
                    f'<span style="background:#DCFCE7;color:#166534;padding:3px 10px;'
                    f'border-radius:12px;font-size:0.82rem;font-weight:600">'
                    f'✅ {_ol_n_listo} listo{"s" if _ol_n_listo != 1 else ""}</span>'
                )
                _ol_badges.append(
                    f'<span style="background:#FEF9C3;color:#854D0E;padding:3px 10px;'
                    f'border-radius:12px;font-size:0.82rem;font-weight:600">'
                    f'⚠ {_ol_n_warn} advertencia{"s" if _ol_n_warn != 1 else ""}</span>'
                )
                _ol_badges.append(
                    f'<span style="background:#FEE2E2;color:#991B1B;padding:3px 10px;'
                    f'border-radius:12px;font-size:0.82rem;font-weight:600">'
                    f'❌ {_ol_n_err} con error</span>'
                )
                st.markdown(
                    '<div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px">'
                    + "".join(_ol_badges) + "</div>",
                    unsafe_allow_html=True,
                )

                # Función de color de filas (idéntica a la del módulo Archivo plano)
                def _style_ol(row: pd.Series) -> list[str]:
                    estado = row["Estado"]
                    resp   = str(row.get("RespuestaAPI", ""))
                    if estado == "❌ Error":
                        color = "background-color: #FEE2E2"
                    elif estado == "✅ Cargado":
                        color = "background-color: #DCFCE7"
                    elif "⚠" in resp:
                        color = "background-color: #FEF9C3"
                    else:
                        color = ""
                    return [color] * len(row)

                # Configuración de columnas reutilizable
                _ol_col_cfg = {
                    "Proyecto":      st.column_config.TextColumn("Proyecto"),
                    "GrupoTarea":    st.column_config.TextColumn("Grupo Tarea"),
                    "Tarea":         st.column_config.TextColumn("Tarea"),
                    "TipoHora":      st.column_config.SelectboxColumn(
                                         "Tipo Hora", options=["Laboral", "Adicional"]),
                    "FechaRegistro": st.column_config.TextColumn("Fecha (YYYY-MM-DD)"),
                    "Horas":         st.column_config.NumberColumn(
                                         "Horas", format="%d h", min_value=0),
                    "Minutos":       st.column_config.SelectboxColumn(
                                         "Minutos", options=[0, 15, 30, 45]),
                    "HoraInicio":    st.column_config.TextColumn("Hora Inicio"),
                    "Comentario":    st.column_config.TextColumn("Comentario"),
                    "Estado":        st.column_config.TextColumn("Estado"),
                    "RespuestaAPI":  st.column_config.TextColumn("Respuesta / Advertencia"),
                }

                _ol_upload_result = st.session_state.get("outlook__upload_result")

                # ── Vista resultados (post-upload) ────────────────────────────
                if _ol_upload_result is not None:
                    _ol_df_res    = _ol_upload_result
                    _ol_res_ok    = int((_ol_df_res["Estado"] == "✅ Cargado").sum())
                    _ol_res_fail  = int((_ol_df_res["Estado"] == "❌ Error").sum())

                    if _ol_res_fail == 0:
                        st.success(
                            f"🎉 ¡Todo listo! **{_ol_res_ok}** "
                            f"registro{'s' if _ol_res_ok != 1 else ''} "
                            f"enviado{'s' if _ol_res_ok != 1 else ''} correctamente a BISOPI."
                        )
                    else:
                        st.warning(
                            f"Upload completado con resultados mixtos — "
                            f"**{_ol_res_ok}** exitoso{'s' if _ol_res_ok != 1 else ''}, "
                            f"**{_ol_res_fail}** fallido{'s' if _ol_res_fail != 1 else ''}."
                        )

                    if _LOCAL_MODE and st.session_state.get("outlook__write_ok"):
                        st.success("💾 Histórico actualizado en la plantilla.")

                    _ol_res_display = _ol_df_res[
                        ["Estado"] + [c for c in _ol_df_res.columns if c != "Estado"]
                    ]
                    _ol_edited_res = st.data_editor(
                        _ol_res_display.style.apply(_style_ol, axis=1),
                        use_container_width=True,
                        hide_index=True,
                        disabled=["Estado", "RespuestaAPI"],
                        column_config=_ol_col_cfg,
                    )

                    # Persistir ediciones para reintento
                    if len(_ol_edited_res) == len(_ol_df_res):
                        _ol_updated = _ol_df_res.copy()
                        for _col in [c for c in _ol_edited_res.columns
                                     if c not in ("Estado", "RespuestaAPI")]:
                            if _col in _ol_updated.columns:
                                _ol_updated[_col] = _ol_edited_res[_col].values
                        st.session_state["outlook__df_edit"] = _ol_updated

                    st.markdown("<br>", unsafe_allow_html=True)
                    _ol_act_cols = st.columns([2, 2, 6])

                    with _ol_act_cols[0]:
                        if _ol_res_fail > 0 and st.button(
                            f"🔄 Reintentar fallidos ({_ol_res_fail})",
                            use_container_width=True,
                            key="outlook__btn_retry",
                        ):
                            _ol_retry = st.session_state.get(
                                "outlook__df_edit", _ol_df_res
                            ).copy()
                            _ol_mask_err = _ol_retry["Estado"] == "❌ Error"
                            _ol_retry.loc[_ol_mask_err, "Estado"]       = "Pendiente"
                            _ol_retry.loc[_ol_mask_err, "RespuestaAPI"] = ""
                            st.session_state["outlook__df_edit"]        = _ol_retry
                            st.session_state["outlook__upload_result"]  = None
                            st.session_state["outlook__write_ok"]       = None
                            st.session_state["outlook__download_bytes"] = None
                            st.rerun()

                    with _ol_act_cols[1]:
                        if not _LOCAL_MODE:
                            _ol_dl = st.session_state.get("outlook__download_bytes")
                            if _ol_dl:
                                st.download_button(
                                    label="⬇ Descargar Histórico",
                                    data=_ol_dl,
                                    file_name="BISOPI_Historico_Outlook.xlsx",
                                    mime=(
                                        "application/vnd.openxmlformats-"
                                        "officedocument.spreadsheetml.sheet"
                                    ),
                                    use_container_width=True,
                                    help="Archivo Excel con el Histórico de los registros enviados.",
                                )

                # ── Vista edición (pre-upload) ────────────────────────────────
                else:
                    _ol_display = _ol_validado[
                        ["Estado"] + [c for c in _ol_validado.columns if c != "Estado"]
                    ]
                    _ol_edited = st.data_editor(
                        _ol_display.style.apply(_style_ol, axis=1),
                        use_container_width=True,
                        hide_index=True,
                        num_rows="dynamic",
                        disabled=["Estado", "RespuestaAPI"],
                        column_config=_ol_col_cfg,
                    )

                    # Persistir ediciones en session_state
                    _ol_new_edit = _ol_edited[_EDITABLE_COLS].copy()
                    _ol_old_edit = st.session_state["outlook__df_edit"]
                    _ol_old_cols = [c for c in _EDITABLE_COLS if c in _ol_old_edit.columns]

                    if "Estado" in _ol_edited.columns:
                        _ol_new_edit["Estado"] = _ol_edited["Estado"].values
                    if "RespuestaAPI" in _ol_edited.columns:
                        _ol_new_edit["RespuestaAPI"] = _ol_edited["RespuestaAPI"].values

                    if len(_ol_new_edit) != len(_ol_old_edit):
                        _ol_changed = True
                    else:
                        _ol_changed = (
                            _ol_new_edit[_ol_old_cols].astype(str).values.tolist()
                            != _ol_old_edit[_ol_old_cols].astype(str).values.tolist()
                        )

                    st.session_state["outlook__df_edit"] = _ol_new_edit.reset_index(drop=True)

                    if _ol_changed:
                        st.rerun()

                    # ── Botones: exportar y subir ─────────────────────────────
                    st.markdown("<br>", unsafe_allow_html=True)

                    _ol_col_export, _ol_col_upload = st.columns(2)

                    # ── Columna izquierda: Exportar a plantilla Excel ─────────
                    with _ol_col_export:
                        if not PLANTILLA_PATH:
                            st.warning(
                                "Configura `PLANTILLA_PATH` en tu `.env` "
                                "para usar esta función."
                            )
                        else:
                            try:
                                _ol_exp_bytes = export_to_template(
                                    to_bisopi_df(st.session_state["outlook__df_edit"])
                                )
                                st.download_button(
                                    label="📥 Exportar a plantilla Excel",
                                    data=_ol_exp_bytes,
                                    file_name="BISOPI_Outlook_Export.xlsx",
                                    mime=(
                                        "application/vnd.openxmlformats-"
                                        "officedocument.spreadsheetml.sheet"
                                    ),
                                    use_container_width=True,
                                    help=(
                                        "Exporta los eventos a la plantilla Excel "
                                        "para completar campos faltantes."
                                    ),
                                )
                                st.caption(
                                    "💡 Abre el archivo, completa los campos "
                                    "resaltados en amarillo (Proyecto, Grupo Tarea, "
                                    "Tarea, Tipo Hora) y cárgalo en la pestaña "
                                    "**Archivo plano**."
                                )
                            except ValueError as _exc:
                                st.warning(str(_exc))
                            except Exception as _exc:
                                st.error(f"Error al generar el archivo: {_exc}")

                    # ── Columna derecha: Subir a BISOPI ───────────────────────
                    with _ol_col_upload:
                        if _ol_n_subir > 0:
                            _ol_btn_lbl = (
                                f"🚀 Subir a BISOPI  "
                                f"({_ol_n_subir} registro"
                                f"{'s' if _ol_n_subir != 1 else ''})"
                            )
                            if st.button(
                                _ol_btn_lbl,
                                type="primary",
                                use_container_width=True,
                                key="outlook__btn_upload",
                            ):
                                # Contar filas que no se enviarán (no Pendiente, no Error)
                                _ol_n_skip = int(
                                    st.session_state["outlook__df_edit"]["Estado"].isin(
                                        {"⚠ Sin clasificar", "⚠ Incompleto"}
                                    ).sum()
                                )
                                if _ol_n_skip > 0:
                                    st.session_state["outlook__upload_confirm"] = True
                                else:
                                    st.session_state["outlook__upload_confirm"] = "go"
                                st.rerun()
                        else:
                            st.info(
                                "No hay registros listos para enviar. "
                                "Corrige los errores marcados en rojo antes de continuar."
                            )

                    # ── Diálogo de confirmación ───────────────────────────────
                    if st.session_state.get("outlook__upload_confirm") is True:
                        _ol_n_skip = int(
                            st.session_state["outlook__df_edit"]["Estado"].isin(
                                {"⚠ Sin clasificar", "⚠ Incompleto"}
                            ).sum()
                        )
                        st.warning(
                            f"⚠️ Hay **{_ol_n_skip}** "
                            f"fila{'s' if _ol_n_skip != 1 else ''} sin clasificar "
                            f"o incompleta{'s' if _ol_n_skip != 1 else ''} que no se "
                            f"enviarán. ¿Deseas continuar solo con las filas completas?"
                        )
                        _ol_cc, _ol_cx, _ = st.columns([2, 2, 6])
                        with _ol_cc:
                            if st.button(
                                "✅ Confirmar y subir",
                                type="primary",
                                use_container_width=True,
                                key="outlook__btn_confirm",
                            ):
                                st.session_state["outlook__upload_confirm"] = "go"
                                st.rerun()
                        with _ol_cx:
                            if st.button(
                                "✖ Cancelar",
                                use_container_width=True,
                                key="outlook__btn_cancel_confirm",
                            ):
                                st.session_state["outlook__upload_confirm"] = False
                                st.rerun()

                    # ── Ejecución del upload ──────────────────────────────────
                    if st.session_state.get("outlook__upload_confirm") == "go":
                        st.session_state["outlook__upload_confirm"] = False

                        _ol_prog_bar  = st.progress(0.0)
                        _ol_prog_text = st.empty()

                        def _ol_on_progress(current: int, total: int, label: str) -> None:
                            _ol_prog_bar.progress(current / total if total > 0 else 0.0)
                            _ol_prog_text.caption(
                                f"Enviando {current} de {total}: *{label}*"
                            )

                        try:
                            if _LOCAL_MODE and not check_writable(PLANTILLA_PATH):
                                st.warning(
                                    "⚠️ **El archivo está abierto en Excel u otro programa.** "
                                    "Ciérralo y vuelve a intentarlo."
                                )
                                st.stop()

                            # Normalizar + filtrar solo Pendiente
                            _ol_df_clean     = to_bisopi_df(st.session_state["outlook__df_edit"])
                            _ol_df_to_upload = _ol_df_clean[
                                _ol_df_clean["Estado"] == "Pendiente"
                            ].copy()

                            _ol_df_result = upload(
                                _ol_df_to_upload,
                                email_colaborador,
                                BISOPI_API_TOKEN,
                                _ol_on_progress,
                            )
                            _ol_prog_bar.progress(1.0)
                            _ol_prog_text.caption("✅ Proceso de envío completado.")

                            # Merge de resultados por índice al DataFrame completo
                            st.session_state["outlook__df_edit"].update(_ol_df_result)

                            # Filas recién exitosas en este batch
                            _ol_df_new_ok = st.session_state["outlook__df_edit"].loc[
                                _ol_df_to_upload.index[
                                    st.session_state["outlook__df_edit"].loc[
                                        _ol_df_to_upload.index, "Estado"
                                    ] == "✅ Cargado"
                                ]
                            ]
                            _ol_df_full = st.session_state["outlook__df_edit"].copy()

                            # Construir workbook para el Histórico
                            if _LOCAL_MODE:
                                _ol_wb = load_workbook(
                                    PLANTILLA_PATH, keep_vba=False, data_only=False
                                )
                            else:
                                _ol_dl_prev = st.session_state.get("outlook__download_bytes")
                                if _ol_dl_prev:
                                    _ol_wb = load_workbook(
                                        io.BytesIO(_ol_dl_prev),
                                        keep_vba=False, data_only=False,
                                    )
                                else:
                                    _ol_wb = Workbook()
                                    _ol_default = _ol_wb.active
                                    if _ol_default and _ol_default.title in ("Sheet", "Sheet1"):
                                        _ol_wb.remove(_ol_default)

                            if not _ol_df_new_ok.empty:
                                append_historico(_ol_wb, _ol_df_new_ok)

                            _ol_eff_path = PLANTILLA_PATH if _LOCAL_MODE else None
                            _ol_write_ok, _ol_dl_bytes = save_plantilla(
                                _ol_wb, _ol_eff_path
                            )

                            st.session_state["outlook__upload_result"]  = _ol_df_full
                            st.session_state["outlook__write_ok"]       = _ol_write_ok
                            st.session_state["outlook__download_bytes"] = _ol_dl_bytes
                            st.rerun()

                        except TokenError:
                            st.error(
                                "❌ **Token inválido (401 Unauthorized).** "
                                "Verifica el valor de **BISOPI_API_TOKEN** "
                                "en tu archivo `.env` y reinicia la aplicación."
                            )
                            st.stop()

                # ── Resumen de la semana ──────────────────────────────────────
                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown("#### Resumen de la semana")

                _ol_smr = build_summary(st.session_state["outlook__df_edit"].copy())

                _ol_c1, _ol_c2, _ol_c3, _ol_c4 = st.columns(4)
                _ol_c1.metric("Registros",          _ol_smr["total_registros"])
                _ol_c2.metric("Horas laborales",    _ol_smr["horas_laborales_fmt"])
                _ol_c3.metric("Horas adicionales",  _ol_smr["horas_adicionales_fmt"])
                _ol_c4.metric("Días con registros", _ol_smr["dias_distintos"])

                st.markdown("<br>", unsafe_allow_html=True)

                def _style_ol_resumen(df: pd.DataFrame) -> pd.DataFrame:
                    styles = pd.DataFrame("", index=df.index, columns=df.columns)
                    for i, mins in enumerate(_ol_smr["laboral_mins_por_dia"]):
                        if mins > 480:
                            styles.loc[i, "Horas Laborales"] = (
                                "background-color: #FEF9C3; color: #854D0E; font-weight: 600"
                            )
                    styles.iloc[-1] = "font-weight: bold; background-color: #F1F5F9"
                    return styles

                st.dataframe(
                    _ol_smr["resumen_df"].style.apply(_style_ol_resumen, axis=None),
                    use_container_width=True,
                    hide_index=True,
                )

                # ── Advertencia de horas laborales faltantes ──────────────────
                _ol_gaps = calculate_gaps(
                    st.session_state["outlook__df_edit"].copy(), work_hours_per_day=8
                )
                if not _ol_gaps.empty:
                    _gap_lines: list[str] = []
                    for _, _grow in _ol_gaps.iterrows():
                        _gh, _gm = divmod(int(_grow["MinutosFaltantes"]), 60)
                        _gap_lines.append(
                            f"- **{_grow['FechaRegistro']}** ({_grow['DiaSemana']}): "
                            f"faltan **{_gh}h {_gm:02d}m** para completar 8h laborales"
                        )
                    st.info(
                        "⏱ Días con horas laborales incompletas:\n\n"
                        + "\n".join(_gap_lines)
                    )

    _render_ol_results()


# ── Tab 4: Acerca de ──────────────────────────────────────────────────────────
with tab_about:
    st.markdown("<br>", unsafe_allow_html=True)

    # ── Bloque 1 — ¿Qué es BISOPI Automator? ─────────────────────────────────
    st.header("¿Qué es BISOPI Automator?")
    st.markdown("""
BISOPI Automator elimina la carga manual de horas en BISOPI: en lugar de ingresar registro
por registro en el portal, preparas tu semana desde una fuente de datos y la app envía todo
en un solo paso con validación previa de reglas de negocio.

Dispone de **tres modos de carga**:

- 📂 **Plantilla Excel** — completa la hoja *Registro* de la plantilla oficial y cárgala directamente, o pega el rango copiado desde Excel.
- 📅 **Agenda Outlook · archivo ICS** — exporta tu calendario como `.ics` y la app extrae automáticamente los eventos marcados para BISOPI.
- 🔗 **Agenda Outlook · Microsoft Graph API** — conecta tu cuenta de Microsoft 365 directamente, sin exportar ningún archivo, mediante autenticación segura con tu cuenta corporativa.
    """)

    st.divider()

    # ── Bloque 2 — Plantilla Excel ────────────────────────────────────────────
    st.header("📂 Cómo usar la plantilla Excel")
    st.markdown("""
1. Descarga la plantilla **`Plantilla_BISOPI_Automator.xlsx`** desde el repositorio o solicítala al equipo.
2. Abre el archivo y dirígete a la hoja **Registro** — ingresa cada registro a partir de la fila 3, una fila por hora imputada.
3. Usa los desplegables de **Tipo Hora** (`Laboral` / `Adicional`) y **Minutos** (`0` / `15` / `30` / `45`).
4. Para horas **Adicionales** completa también **Hora Inicio** en formato `HH:MM` y el campo **Comentario** (ambos obligatorios).
5. Los valores de **Proyecto**, **Grupo Tarea** y **Tarea** deben coincidir **exactamente** con los nombres en BISOPI — incluyendo mayúsculas, tildes y espacios.
6. Consulta la hoja **Catálogos** del mismo archivo para copiar los nombres correctos sin riesgo de error tipográfico.
7. Consulta la hoja **Ejemplos** para ver casos representativos de distintos tipos de registro.
    """)

    st.divider()

    # ── Bloque 3 — Integración con Outlook ───────────────────────────────────
    st.header("📅 Integración con Outlook")

    st.subheader("Carga por archivo ICS")
    st.markdown("""
1. **Exporta tu calendario** desde Outlook como archivo `.ics`:
   - **Outlook Web:** Ve a *Calendario → Configuración (⚙️) → Calendarios compartidos → Publicar un calendario*. Elige el calendario, define los permisos y haz clic en **Publicar**. Copia el enlace `.ics` generado y ábrelo en el navegador para descargarlo.
   - **Outlook Escritorio (Windows):** *Archivo → Guardar calendario* → elige nombre y haz clic en **Más opciones** para seleccionar el rango de fechas → **Aceptar → Guardar**.
2. En la pestaña **Agenda Outlook**, selecciona el submodo **Archivo ICS** y sube el archivo.
3. Ajusta el **rango de fechas** (inicio y fin de semana) — la app filtra automáticamente los eventos del periodo seleccionado.
4. Revisa los eventos detectados en la tabla editable, completa los campos faltantes y pulsa **Subir a BISOPI**.
    """)

    st.subheader("Convención de eventos para BISOPI")
    st.markdown(
        "Para que la app reconozca un evento automáticamente, el **título** debe comenzar con `[BISOPI]`:\n\n"
        "```\n[BISOPI] Revisión de sprint con el equipo\n```\n\n"
        "La **descripción** del evento puede incluir los siguientes campos (uno por línea):"
    )
    st.code(
        "proyecto: NOMBRE-EXACTO-DEL-PROYECTO\n"
        "grupo:    Nombre del grupo de tarea\n"
        "tarea:    Nombre exacto de la tarea\n"
        "tipo:     Laboral\n"
        "comentario: Descripción opcional",
        language="text",
    )
    st.markdown("""
> ⚠️ **Nota:** Los eventos **sin** `[BISOPI]` en el título se importan igualmente como **"sin clasificar"**.
> Puedes completar los campos directamente en la tabla editable de la app antes de enviar — ningún evento se descarta automáticamente.
    """)

    st.subheader("Conexión directa con Microsoft 365 (Graph API)")
    st.markdown("""
Requiere que IT registre la aplicación en Azure AD. El flujo de autenticación varía
según el modo de distribución:

**☁️ Versión Cloud:**
La autenticación ocurre en la **pantalla de inicio** de la app, antes de ver cualquier pestaña.
Al autenticarte, tu email corporativo se obtiene automáticamente y queda disponible
para toda la sesión — incluyendo la pestaña Agenda Outlook, que ya aparece conectada sin pasos adicionales.

**💻 Versión Local:**
La autenticación se realiza desde la pestaña **Agenda Outlook** al seleccionar el submodo
**Microsoft Graph API**.

En ambos casos puedes elegir entre dos métodos:

- **🖥️ Ventana de login** — abre el navegador para iniciar sesión con tu cuenta corporativa.
  Requiere que IT registre la Redirect URI de la app en Azure AD.
- **📱 Código de dispositivo** — sin ventanas emergentes; te da un código para ingresar en
  `microsoft.com/devicelogin` desde cualquier dispositivo. No requiere configuración adicional en Azure AD.

Una vez conectado, selecciona el rango de fechas en la pestaña **Agenda Outlook** y pulsa
**Obtener eventos del calendario**.
    """)

    st.divider()

    # ── Bloque 4 — Reglas de negocio ─────────────────────────────────────────
    st.header("📋 Reglas de negocio")
    _rules_df = pd.DataFrame({
        "Regla": [
            "Límite semanal",
            "Semana cerrada",
            "Horas adicionales",
            "Medianoche",
            "Prioridad de envío",
        ],
        "Detalle": [
            "Máximo 40 horas laborales por semana (lunes a domingo).",
            "Los lunes a las 12:30 se cierra la semana anterior. Contacta a la persona responsable para abrirla excepcionalmente.",
            "Requieren Hora Inicio (HH:MM) y Comentario obligatorios.",
            "Una hora adicional no puede superar las 23:59 del mismo día.",
            "Las horas laborales se envían antes que las adicionales automáticamente.",
        ],
    })
    st.table(_rules_df)

    st.divider()

    # ── Bloque 5 — ¿Qué hacer si un registro falla? ──────────────────────────
    st.header("🔴 ¿Qué hacer si un registro falla?")
    st.markdown("""
- Los registros fallidos quedan **resaltados en rojo** con el mensaje de error exacto devuelto por la API.
- Corrige los datos directamente en la **tabla editable** — sin necesidad de recargar el archivo.
- Pulsa **"Reintentar fallidos"** — solo se reenvían los registros que fallaron, no los ya exitosos.
- Si el error indica **"semana cerrada"**: contacta a la persona responsable para solicitar la apertura excepcional.
- Si el error indica **"proyecto no encontrado"**: verifica que el nombre coincida exactamente con BISOPI (mayúsculas, tildes y espacios incluidos).
    """)

    st.divider()

    # ── Bloque 6 — Modos de distribución ─────────────────────────────────────
    st.header("☁️ Modos de distribución")
    st.markdown("""
BISOPI Automator puede usarse de dos formas. El comportamiento cambia automáticamente
según la variable de entorno `BISOPI_ENV`.
    """)
    _dist_df = pd.DataFrame({
        "Característica": [
            "Acceso",
            "Instalación",
            "Email del colaborador",
            "Plantilla Excel",
            "Autenticación Outlook",
            "Actualizaciones",
            "Requiere internet",
        ],
        "☁️ Cloud": [
            "URL desde cualquier navegador",
            "Ninguna — solo abrir el link",
            "Automático desde el login con Microsoft",
            "Se descarga al final de cada sesión",
            "En la pantalla de inicio, antes de entrar a la app",
            "Automáticas con cada cambio en el repositorio",
            "Sí",
        ],
        "💻 Local": [
            "Solo el equipo donde está instalado",
            "Python + install.bat (una sola vez)",
            "Configurado en el archivo .env",
            "Lectura y escritura directa en disco",
            "En la pestaña Agenda Outlook",
            "Manuales — reemplazar archivos",
            "Solo para API de BISOPI y Outlook",
        ],
    })
    st.table(_dist_df)

    st.divider()

    # ── Bloque 7 — Créditos ──────────────────────────────────────────────────
    _env_label = "☁️ Cloud" if is_cloud() else "💻 Local"
    st.markdown(
        "<div style='color:#94a3b8;font-size:0.72rem;font-weight:600;letter-spacing:1.5px;"
        "text-transform:uppercase;margin-bottom:4px;'>Créditos</div>"
        f"<div style='color:#64748b;font-size:0.8rem;margin-bottom:14px;'>"
        f"v1.0 — Mayo 2026 &nbsp;·&nbsp; {_env_label}</div>",
        unsafe_allow_html=True,
    )

    # Dos columnas: Autor | Stack técnico
    col_autor, col_stack = st.columns(2, gap="medium")

    with col_autor:
        st.markdown("""
        <div style="background:#f8fafc;border:1.5px solid #e2e8f0;border-radius:10px;
                    padding:20px 22px;height:100%;">
          <div style="color:#64748b;font-size:0.72rem;font-weight:600;letter-spacing:1px;
                      text-transform:uppercase;margin-bottom:10px;">Autor</div>
          <div style="color:#1e293b;font-size:1rem;font-weight:600;">Edson Leon</div>
          <div style="color:#475569;font-size:0.85rem;margin-top:4px;">
            Data Insights Lead — Operaciones
          </div>
          <div style="color:#475569;font-size:0.85rem;">Bision Consulting</div>
          <div style="margin-top:10px;">
            <a href="mailto:edson.leon@bisionconsulting.com"
               style="color:#2563eb;font-size:0.82rem;text-decoration:none;">
              ✉ edson.leon@bisionconsulting.com
            </a>
          </div>
        </div>
        """, unsafe_allow_html=True)

    with col_stack:
        st.markdown("""
        <div style="background:#f8fafc;border:1.5px solid #e2e8f0;border-radius:10px;
                    padding:20px 22px;height:100%;">
          <div style="color:#64748b;font-size:0.72rem;font-weight:600;letter-spacing:1px;
                      text-transform:uppercase;margin-bottom:10px;">Stack técnico</div>
          <div style="display:flex;flex-wrap:wrap;gap:6px;margin-top:4px;">
            <span style="background:#dbeafe;color:#1d4ed8;font-size:0.75rem;font-weight:600;
                         padding:3px 10px;border-radius:20px;">Python</span>
            <span style="background:#dbeafe;color:#1d4ed8;font-size:0.75rem;font-weight:600;
                         padding:3px 10px;border-radius:20px;">Streamlit</span>
            <span style="background:#dbeafe;color:#1d4ed8;font-size:0.75rem;font-weight:600;
                         padding:3px 10px;border-radius:20px;">pandas</span>
            <span style="background:#dbeafe;color:#1d4ed8;font-size:0.75rem;font-weight:600;
                         padding:3px 10px;border-radius:20px;">openpyxl</span>
            <span style="background:#dbeafe;color:#1d4ed8;font-size:0.75rem;font-weight:600;
                         padding:3px 10px;border-radius:20px;">msal</span>
            <span style="background:#dbeafe;color:#1d4ed8;font-size:0.75rem;font-weight:600;
                         padding:3px 10px;border-radius:20px;">requests</span>
          </div>
        </div>
        """, unsafe_allow_html=True)

    # Descripción
    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
    st.markdown("""
    <div style="background:#f8fafc;border:1.5px solid #e2e8f0;border-radius:10px;
                padding:22px 24px;">
      <div style="color:#64748b;font-size:0.72rem;font-weight:600;letter-spacing:1px;
                  text-transform:uppercase;margin-bottom:10px;">Descripción</div>
      <p style="color:#334155;font-size:0.88rem;line-height:1.7;margin:0;">
        Herramienta de automatización para el registro de horas en BISOPI.
        Desarrollada para eliminar el proceso manual de imputación tarea por tarea,
        permitiendo cargar una semana completa desde plantilla Excel, archivo ICS o
        conexión directa con Microsoft 365, con validación previa de reglas de negocio.
        Arquitectura escalable hacia integración con Azure DevOps.
      </p>
    </div>
    """, unsafe_allow_html=True)


# ── Pie de página ─────────────────────────────────────────────────────────────
st.divider()
st.caption(
    "BISOPI Automator v1.0 — Mayo 2026 · "
    "Desarrollado por Edson Leon · "
    "Bision Consulting"
)
